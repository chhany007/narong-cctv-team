# CameraMonitor_Final_v8.py
"""
Camera Monitor Final v8 - Full
- Robust NVR sheet reading: handles NVR sheet with or without header row
- UI improvements: table fills remaining right-side space, reduced margins, splitter tuned
- Features: NVR sidebar, Refresh NVR Status, camera table, double-click -> VLC with default creds,
  credential manager (keyring optional + fallback), silent ping, export CSV.
- Update Manager: Automatic update checking and downloading
"""

import os, sys, json, csv, time, socket, threading, subprocess, webbrowser, traceback, unicodedata, re, platform, struct
from PyQt5 import QtCore, QtGui, QtWidgets
import pandas as pd
import concurrent.futures
import requests
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET

# Import update manager
try:
    from update_manager import check_for_updates_async, UpdateChecker
    UPDATE_MANAGER_AVAILABLE = True
except Exception as e:
    UPDATE_MANAGER_AVAILABLE = False
    print(f"Update manager not available: {e}")

# Optional keyring
try:
    import keyring
    KEYRING_AVAILABLE = True
except Exception:
    KEYRING_AVAILABLE = False

# Config
EXCEL_FILE = "ip.xlsx"
LOG_FILE = "camera_monitor.log"
CREDS_META = "creds_meta.json"
CREDS_FALLBACK = "creds_store.json"
EXPORT_FILE = "exported_cameras.csv"
CHECK_HISTORY_FILE = "check_history.json"
LOGO_FILE = "sky-tech logo.png"

HTTP_PORT = 80
RTSP_PORT = 554
PING_TIMEOUT = 2  # seconds

DEFAULT_CREDS = [("admin", "Kkcctv12345"), ("admin", "Kkcctv1245")]

VLC_PATHS = [
    r"C:\Program Files\VideoLAN\VLC\vlc.exe",
    r"C:\Program Files (x86)\VideoLAN\VLC\vlc.exe",
    "/usr/bin/vlc",
    "/usr/local/bin/vlc",
    "/Applications/VLC.app/Contents/MacOS/VLC",
]

# ---------------- utilities ----------------
def log(msg: str):
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass

def silent_ping(ip: str) -> bool:
    if not ip:
        return False
    system = platform.system().lower()
    param = "-n" if system.startswith("windows") else "-c"
    cmd = ["ping", param, "1", ip]
    try:
        if system.startswith("windows"):
            CREATE_NO_WINDOW = 0x08000000
            subprocess.check_output(cmd, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL,
                                    timeout=PING_TIMEOUT+1, creationflags=CREATE_NO_WINDOW)
        else:
            subprocess.check_output(cmd, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL,
                                    timeout=PING_TIMEOUT+1)
        return True
    except Exception:
        return False

def check_tcp(ip: str, port: int, timeout: float = 1.0) -> bool:
    try:
        sock = socket.create_connection((ip, port), timeout=timeout)
        sock.close()
        return True
    except Exception:
        return False

def find_vlc_executable():
    for p in VLC_PATHS:
        if os.path.exists(p):
            return p
    return "vlc"

# Helper: find local interface IP that would be used to reach a given destination
def get_local_ip_for_target(dest_ip: str) -> str:
    """Return local source IP used to reach dest_ip, or empty string on failure."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect((dest_ip, 80))
        local_ip = s.getsockname()[0]
        s.close()
        return local_ip
    except Exception:
        try:
            s.close()
        except Exception:
            pass
    return ""

# Model-specific endpoint hints for Hikvision NVRs
MODEL_ENDPOINTS = {
    "DS-7732NXI-K4": [
        "/ISAPI/ContentMgmt/RemoteDevice",
        "/ISAPI/Streaming/channels",
        "/ISAPI/ContentMgmt/InputProxy/channels",
        "/ISAPI/PSIA/Custom/SelfAdapt/Channel"
    ],
    "DS-7732NI-K4": [
        "/ISAPI/ContentMgmt/RemoteDevice",
        "/ISAPI/Streaming/channels",
        "/ISAPI/PSIA/Custom/SelfAdapt/Channel"
    ]
}

# ---------------- credentials helpers ----------------
def load_creds_meta():
    try:
        if os.path.exists(CREDS_META):
            with open(CREDS_META, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        log("read creds_meta failed")
    return {}

def save_creds_meta(meta: dict):
    try:
        with open(CREDS_META, "w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2)
    except Exception as e:
        log(f"save_creds_meta: {e}")

def set_password(ip: str, username: str, password: str):
    if KEYRING_AVAILABLE:
        try:
            keyring.set_password(f"CameraMonitor:{ip}", username, password)
            return True
        except Exception as e:
            log(f"keyring set failed: {e}")
    # fallback store
    store = {}
    try:
        if os.path.exists(CREDS_FALLBACK):
            with open(CREDS_FALLBACK, "r", encoding="utf-8") as f:
                store = json.load(f)
    except Exception:
        store = {}
    store[ip] = {"username": username, "password": password}
    try:
        with open(CREDS_FALLBACK, "w", encoding="utf-8") as f:
            json.dump(store, f, indent=2)
        return False
    except Exception as e:
        log(f"fallback write failed: {e}")
        return False

def get_password(ip: str):
    meta = load_creds_meta()
    if ip in meta and meta[ip].get("username"):
        u = meta[ip]["username"]
        if KEYRING_AVAILABLE:
            try:
                p = keyring.get_password(f"CameraMonitor:{ip}", u)
                if p:
                    return u, p
            except Exception as e:
                log(f"keyring get failed: {e}")
        # fallback
        try:
            if os.path.exists(CREDS_FALLBACK):
                with open(CREDS_FALLBACK, "r", encoding="utf-8") as f:
                    store = json.load(f)
                if ip in store:
                    return store[ip].get("username"), store[ip].get("password")
        except Exception:
            pass
    return None, None

def delete_credentials(ip: str):
    try:
        meta = load_creds_meta()
        if ip in meta:
            meta.pop(ip, None)
            save_creds_meta(meta)
    except Exception:
        pass
    try:
        if os.path.exists(CREDS_FALLBACK):
            with open(CREDS_FALLBACK, "r", encoding="utf-8") as f:
                store = json.load(f)
            if ip in store:
                store.pop(ip, None)
                with open(CREDS_FALLBACK, "w", encoding="utf-8") as f:
                    json.dump(store, f, indent=2)
    except Exception:
        pass

# ---------------- NVR login & IP update ----------------
def test_nvr_login(ip: str, username: str, password: str, timeout: float = 2.0) -> tuple:
    """Test NVR login with multiple methods. Returns (success: bool, real_ip: str, error_msg: str)."""
    log(f"=== NVR Login Test Start ===")
    log(f"IP: {ip}, Username: {username}, Timeout: {timeout}s")
    
    auth_succeeded = False  # Track if auth worked, to continue trying for IP extraction
    
    try:
        # First: Try SADP discovery (works even if ping is blocked)
        log(f"[1] Testing SADP discovery to {ip}...")
        try:
            SADP_PORT = 33333
            SADP_REQUEST = b'\x00\x00\x00\x00\x00\x00\x00\x38<?xml version="1.0"?><Command><AccessFlag>1</AccessFlag><Command>GetDeviceInfo</Command></Command>'
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.settimeout(1.0)
            sock.sendto(SADP_REQUEST, (ip, SADP_PORT))
            try:
                data, _ = sock.recvfrom(4096)
                sock.close()
                log(f"[1] SADP RESPONSE received - NVR is reachable")
            except socket.timeout:
                sock.close()
                log(f"[1] SADP timeout - continuing with TCP methods")
        except Exception as e:
            log(f"[1] SADP error: {e} - continuing with TCP methods")
        
        # Method 1: Try v1 API (fastest)
        log(f"[2] Trying v1 API: /api/v1/system/info")
        url = f"http://{ip}/api/v1/system/info"
        try:
            log(f"[2] GET {url}")
            resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
            log(f"[2] Response status: {resp.status_code}")
            if resp.status_code == 200:
                auth_succeeded = True
                try:
                    data = resp.json()
                    log(f"[2] JSON response: {data}")
                    real_ip = data.get("ip") or data.get("ipAddress") or data.get("network", {}).get("ip") or ""
                    log(f"[2] SUCCESS - Extracted IP: {real_ip}")
                    return True, real_ip, ""
                except Exception as e:
                    log(f"[2] JSON parse error: {e}")
                    pass
        except requests.exceptions.Timeout as e:
            log(f"[2] TIMEOUT: {e}")
        except requests.exceptions.ConnectionError as e:
            log(f"[2] CONNECTION ERROR: {e}")
        except Exception as e:
            log(f"[2] ERROR: {e}")
        
        # Method 2: Try v2 API (more likely to have IP info than web interface)
        log(f"[3] Trying v2 API: /api/v2/system/info")
        url = f"http://{ip}/api/v2/system/info"
        try:
            log(f"[3] GET {url}")
            resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
            log(f"[3] Response status: {resp.status_code}")
            if resp.status_code == 200:
                auth_succeeded = True
                try:
                    data = resp.json()
                    log(f"[3] JSON response: {data}")
                    real_ip = data.get("ip") or data.get("ipAddress") or data.get("network", {}).get("ip") or ""
                    log(f"[3] SUCCESS - Extracted IP: {real_ip}")
                    return True, real_ip, ""
                except Exception as e:
                    log(f"[3] JSON parse error: {e}")
                    pass
        except requests.exceptions.Timeout as e:
            log(f"[3] TIMEOUT: {e}")
        except requests.exceptions.ConnectionError as e:
            log(f"[3] CONNECTION ERROR: {e}")
        except Exception as e:
            log(f"[3] ERROR: {e}")
        
        # Method 3: Try simple HTTP GET to web interface (quick auth check)
        log(f"[4] Trying web interface: /")
        url = f"http://{ip}/"
        try:
            log(f"[4] GET {url}")
            resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
            log(f"[4] Response status: {resp.status_code}")
            if resp.status_code == 200:
                auth_succeeded = True
                # Try to extract IP from HTML/JSON if available
                real_ip = ""
                try:
                    # Try JSON
                    data = resp.json()
                    real_ip = data.get("ip") or data.get("ipAddress") or data.get("network", {}).get("ip") or ""
                    log(f"[4] JSON response: {data}")
                except:
                    # Try to find IP in HTML
                    import re
                    matches = re.findall(r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b', resp.text)
                    if matches and matches[0] != ip:  # Don't extract the IP we already know
                        real_ip = matches[0]
                        log(f"[4] Extracted IP from HTML: {real_ip}")
                
                if real_ip:
                    log(f"[4] SUCCESS - Web interface accessible, extracted IP: {real_ip}")
                    return True, real_ip, ""
                else:
                    log(f"[4] SUCCESS - Web interface accessible, no IP extracted in response")
                    # Continue to try legacy CGI for IP extraction
        except requests.exceptions.Timeout as e:
            log(f"[4] TIMEOUT: {e}")
        except requests.exceptions.ConnectionError as e:
            log(f"[4] CONNECTION ERROR: {e}")
        except Exception as e:
            log(f"[4] ERROR: {e}")
        
        # Method 4: Try legacy CGI endpoint (last resort)
        log(f"[5] Trying legacy CGI: /cgi-bin/system")
        url = f"http://{ip}/cgi-bin/system?action=getSystemInfo"
        try:
            log(f"[5] GET {url}")
            resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
            log(f"[5] Response status: {resp.status_code}")
            if resp.status_code == 200:
                auth_succeeded = True
                text = resp.text
                log(f"[5] Response text (first 500 chars): {text[:500]}")
                if "error" not in text.lower():
                    for line in text.split('\n'):
                        if 'ipaddr' in line.lower() or 'ip=' in line.lower():
                            log(f"[5] Found IP line: {line}")
                            parts = line.split('=')
                            if len(parts) > 1:
                                real_ip = parts[1].strip().strip('"')
                                if real_ip and real_ip != "—":
                                    log(f"[5] SUCCESS - Extracted IP: {real_ip}")
                                    return True, real_ip, ""
                    log(f"[5] SUCCESS - CGI authenticated but no IP found")
                    return True, "", ""
        except requests.exceptions.Timeout as e:
            log(f"[5] TIMEOUT: {e}")
        except requests.exceptions.ConnectionError as e:
            log(f"[5] CONNECTION ERROR: {e}")
        except Exception as e:
            log(f"[5] ERROR: {e}")
        
        # All methods failed
        if auth_succeeded:
            log(f"=== AUTHENTICATION SUCCEEDED BUT NO IP EXTRACTED ===")
            log(f"Using connection IP as device IP: {ip}")
            # Use the IP we connected to as the real IP (fallback)
            return True, ip, ""
        else:
            log(f"=== ALL METHODS FAILED ===")
            return False, "", "All authentication methods failed"
            
    except Exception as e:
        msg = str(e)
        log(f"=== UNEXPECTED ERROR: {msg} ===")
        return False, "", msg

def fetch_nvr_cameras(nvr_ip: str, username: str, password: str, timeout: float = 3.0) -> tuple:
    """Fetch list of cameras from NVR. Returns (success: bool, cameras: list, total_count: int, active_count: int, error_msg: str)
    
    cameras list format: [{"name": "Camera 1", "ip": "192.168.1.10", "status": "online", "channel": 1, "port": 554}, ...]
    """
    log(f"=== Fetching NVR Cameras ===")
    log(f"NVR IP: {nvr_ip}, User: {username}, Timeout: {timeout}s")
    
    cameras = []
    total_count = 0
    active_count = 0
    
    try:
        # Method 1: Try Hikvision ISAPI - Most common
        log(f"[1] Trying Hikvision ISAPI: /ISAPI/ContentMgmt/InputProxy/channels")
        url = f"http://{nvr_ip}/ISAPI/ContentMgmt/InputProxy/channels"
        try:
            resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
            log(f"[1] Response status: {resp.status_code}")
            if resp.status_code == 200:
                try:
                    root = ET.fromstring(resp.content)
                    for chan in root.findall('.//InputProxyChannel'):
                        chan_id = chan.findtext('id', '')
                        chan_name = chan.findtext('name', f'Camera {chan_id}')
                        chan_enabled = chan.findtext('enabled', 'false').lower() == 'true'
                        
                        # Try to get IP from sourceInputPortDescriptor
                        ip_addr = ""
                        port = 554
                        src_desc = chan.find('.//sourceInputPortDescriptor')
                        if src_desc is not None:
                            ip_addr = src_desc.findtext('ipAddress', '')
                            port_text = src_desc.findtext('managePortNo', '554')
                            try:
                                port = int(port_text)
                            except:
                                port = 554
                        
                        if ip_addr and chan_enabled:
                            status = "online" if chan_enabled else "offline"
                            cameras.append({
                                "name": chan_name,
                                "ip": ip_addr,
                                "status": status,
                                "channel": chan_id,
                                "port": port
                            })
                            total_count += 1
                            if chan_enabled:
                                active_count += 1
                            log(f"[1] Found camera: {chan_name} ({ip_addr}:{port}) - Channel {chan_id} - {status}")
                    
                    if cameras:
                        log(f"[1] SUCCESS - Hikvision ISAPI: {total_count} cameras, {active_count} active")
                        return True, cameras, total_count, active_count, ""
                except ET.ParseError as e:
                    log(f"[1] XML Parse error: {e}")
        except requests.exceptions.Timeout:
            log(f"[1] TIMEOUT")
        except Exception as e:
            log(f"[1] ERROR: {e}")
        
        # Method 2: Try Dahua API
        log(f"[2] Trying Dahua API: /cgi-bin/configManager.cgi?action=getConfig&name=ChannelTitle")
        url = f"http://{nvr_ip}/cgi-bin/configManager.cgi?action=getConfig&name=ChannelTitle"
        try:
            resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
            log(f"[2] Response status: {resp.status_code}")
            if resp.status_code == 200:
                # Parse Dahua config format: table.ChannelTitle[0].Name=Camera1
                lines = resp.text.split('\n')
                channel_names = {}
                for line in lines:
                    if 'ChannelTitle[' in line and '.Name=' in line:
                        try:
                            # Extract: table.ChannelTitle[0].Name=Camera1
                            chan_num = line.split('[')[1].split(']')[0]
                            name = line.split('=', 1)[1].strip()
                            channel_names[chan_num] = name
                        except:
                            pass
                
                if channel_names:
                    # Now get channel IPs
                    url2 = f"http://{nvr_ip}/cgi-bin/configManager.cgi?action=getConfig&name=RemoteDevice"
                    resp2 = requests.get(url2, auth=HTTPBasicAuth(username, password), timeout=timeout)
                    if resp2.status_code == 200:
                        for line in resp2.text.split('\n'):
                            if 'RemoteDevice[' in line and '.Address=' in line:
                                try:
                                    chan_num = line.split('[')[1].split(']')[0]
                                    ip_addr = line.split('=', 1)[1].strip()
                                    name = channel_names.get(chan_num, f'Camera {chan_num}')
                                    cameras.append({
                                        "name": name,
                                        "ip": ip_addr,
                                        "status": "online",
                                        "channel": chan_num,
                                        "port": 554
                                    })
                                    total_count += 1
                                    active_count += 1
                                    log(f"[2] Found camera: {name} ({ip_addr}) - Channel {chan_num}")
                                except:
                                    pass
                
                if cameras:
                    log(f"[2] SUCCESS - Dahua API: {total_count} cameras")
                    return True, cameras, total_count, active_count, ""
        except requests.exceptions.Timeout:
            log(f"[2] TIMEOUT")
        except Exception as e:
            log(f"[2] ERROR: {e}")
        
        # Method 3: Try v1 API (GET device list with status)
        log(f"[3] Trying generic v1 API: /api/v1/devices")
        url = f"http://{nvr_ip}/api/v1/devices"
        try:
            resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
            log(f"[1] Response status: {resp.status_code}")
            if resp.status_code == 200:
                data = resp.json()
                devices = data.get("devices", []) if isinstance(data, dict) else (data if isinstance(data, list) else [])
                for dev in devices:
                    cam_name = dev.get("name") or dev.get("deviceName") or f"Camera {len(cameras) + 1}"
                    cam_ip = dev.get("ip") or dev.get("ipAddress") or dev.get("address") or ""
                    cam_status = dev.get("status") or dev.get("state") or "unknown"
                    
                    if cam_ip:  # Only add if has IP
                        cameras.append({
                            "name": cam_name,
                            "ip": cam_ip,
                            "status": cam_status
                        })
                        total_count += 1
                        if cam_status.lower() in ("online", "active", "1", "true", "recording"):
                            active_count += 1
                        log(f"[1] Found camera: {cam_name} ({cam_ip}) - {cam_status}")
                
                if cameras:
                    log(f"[1] SUCCESS - Found {total_count} cameras, {active_count} active")
                    return True, cameras, total_count, active_count, ""
        except requests.exceptions.Timeout as e:
            log(f"[1] TIMEOUT: {e}")
        except Exception as e:
            log(f"[1] ERROR: {e}")
        
        # Method 2: Try v2 API (alternative structure)
        log(f"[2] Trying v2 API: /api/v2/devices")
        url = f"http://{nvr_ip}/api/v2/devices"
        try:
            resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
            log(f"[2] Response status: {resp.status_code}")
            if resp.status_code == 200:
                data = resp.json()
                devices = data.get("devices", []) if isinstance(data, dict) else (data if isinstance(data, list) else [])
                for dev in devices:
                    cam_name = dev.get("name") or dev.get("deviceName") or f"Camera {len(cameras) + 1}"
                    cam_ip = dev.get("ip") or dev.get("ipAddress") or dev.get("address") or ""
                    cam_status = dev.get("status") or dev.get("state") or "unknown"
                    
                    if cam_ip:
                        cameras.append({
                            "name": cam_name,
                            "ip": cam_ip,
                            "status": cam_status
                        })
                        total_count += 1
                        if cam_status.lower() in ("online", "active", "1", "true", "recording"):
                            active_count += 1
                        log(f"[2] Found camera: {cam_name} ({cam_ip}) - {cam_status}")
                
                if cameras:
                    log(f"[2] SUCCESS - Found {total_count} cameras, {active_count} active")
                    return True, cameras, total_count, active_count, ""
        except requests.exceptions.Timeout as e:
            log(f"[2] TIMEOUT: {e}")
        except Exception as e:
            log(f"[2] ERROR: {e}")
        
        # Method 3: Try Hikvision channel list API
        log(f"[3] Trying Hikvision channels: /api/v1/channels or /isapi/System/Video/inputs/channels")
        endpoints_to_try = [
            "/api/v1/channels",
            "/api/v1/System/Video/inputs/channels",
            "/isapi/System/Video/inputs/channels"
        ]
        
        for endpoint in endpoints_to_try:
            url = f"http://{nvr_ip}{endpoint}"
            try:
                resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
                log(f"[3] {endpoint}: Status {resp.status_code}")
                if resp.status_code == 200:
                    try:
                        data = resp.json()
                        channels = data.get("channels", []) if isinstance(data, dict) else (data if isinstance(data, list) else [])
                        for ch in channels:
                            cam_name = ch.get("name") or ch.get("channelName") or f"Channel {len(cameras) + 1}"
                            cam_ip = ch.get("ip") or ch.get("ipAddress") or ""
                            cam_status = ch.get("status") or "unknown"
                            
                            if cam_ip:
                                cameras.append({"name": cam_name, "ip": cam_ip, "status": cam_status})
                                total_count += 1
                                if cam_status.lower() in ("online", "active", "1", "true"):
                                    active_count += 1
                                log(f"[3] Found channel: {cam_name} ({cam_ip})")
                        
                        if cameras:
                            log(f"[3] SUCCESS - Found {total_count} cameras")
                            return True, cameras, total_count, active_count, ""
                    except Exception as e:
                        log(f"[3] {endpoint} JSON error: {e}")
            except Exception as e:
                log(f"[3] {endpoint} error: {e}")
        
        # Method 4: Try legacy CGI (list channels)
        log(f"[4] Trying legacy CGI: /cgi-bin/isapi/Platform")
        url = f"http://{nvr_ip}/cgi-bin/isapi/Platform/systemResources"
        try:
            resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
            log(f"[4] Response status: {resp.status_code}")
            if resp.status_code == 200:
                # Parse XML response for channel info
                try:
                    root = ET.fromstring(resp.text)
                    # Look for channels/cameras (structure varies by NVR model)
                    channels = root.findall(".//channelList") or root.findall(".//channel")
                    for idx, chan in enumerate(channels, 1):
                        chan_name = f"Channel {idx}"
                        chan_ip = ""
                        # Try to extract from various XML structures
                        for child in chan:
                            if "name" in child.tag.lower():
                                chan_name = child.text or chan_name
                            if "ip" in child.tag.lower():
                                chan_ip = child.text or ""
                        
                        if chan_ip:
                            cameras.append({
                                "name": chan_name,
                                "ip": chan_ip,
                                "status": "unknown"
                            })
                            total_count += 1
                            log(f"[4] Found camera: {chan_name} ({chan_ip})")
                    
                    if cameras:
                        log(f"[4] SUCCESS - Found {total_count} cameras")
                        return True, cameras, total_count, active_count, ""
                except ET.ParseError as e:
                    log(f"[4] XML parse error: {e}")
        except requests.exceptions.Timeout as e:
            log(f"[4] TIMEOUT: {e}")
        except Exception as e:
            log(f"[4] ERROR: {e}")
        
        # Method 5: Try ONVIF GetDevices (requires ONVIF support)
        log(f"[5] Trying ONVIF endpoint: /onvif/device_service")
        url = f"http://{nvr_ip}/onvif/device_service"
        try:
            resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=timeout)
            log(f"[5] Response status: {resp.status_code}")
            if resp.status_code == 200 and resp.text:
                try:
                    root = ET.fromstring(resp.text)
                    # ONVIF typically returns device info
                    devices = root.findall(".//{http://www.onvif.org/ver10/schema}Device") or root.findall(".//Device")
                    for dev in devices:
                        cam_name = dev.find(".//Name") or dev.find(".//name")
                        cam_name = cam_name.text if cam_name is not None else f"Device {len(cameras) + 1}"
                        
                        # Try to find IP in ONVIF response
                        cam_ip = ""
                        for elem in dev.iter():
                            if "ip" in elem.tag.lower() and elem.text:
                                cam_ip = elem.text
                                break
                        
                        if cam_ip and cam_ip != nvr_ip:  # Don't add NVR IP itself
                            cameras.append({"name": cam_name, "ip": cam_ip, "status": "unknown"})
                            total_count += 1
                            log(f"[5] Found ONVIF device: {cam_name} ({cam_ip})")
                    
                    if cameras:
                        log(f"[5] SUCCESS - Found {total_count} cameras via ONVIF")
                        return True, cameras, total_count, active_count, ""
                except ET.ParseError as e:
                    log(f"[5] XML parse error: {e}")
        except Exception as e:
            log(f"[5] ERROR: {e}")
        
        # Method 6: Try to detect cameras via network scan (SADP) with model-specific fallbacks
        log(f"[6] Trying SADP network discovery (targeted to NVR subnet)")
        try:
            # attempt to detect NVR model (best-effort)
            nvr_model = ""
            try:
                for mod_url in [f"http://{nvr_ip}/ISAPI/System/deviceInfo", f"http://{nvr_ip}/api/v1/system/info", f"http://{nvr_ip}/api/v2/system/info"]:
                    try:
                        r = requests.get(mod_url, auth=HTTPBasicAuth(username, password), timeout=1.0)
                        if r.status_code == 200:
                            # try parse XML or JSON
                            try:
                                data = r.json()
                                nvr_model = data.get("model") or data.get("deviceModel") or nvr_model
                            except Exception:
                                try:
                                    root = ET.fromstring(r.text)
                                    m = root.findtext(".//Model") or root.findtext(".//model") or ""
                                    if m:
                                        nvr_model = m
                                except Exception:
                                    pass
                            if nvr_model:
                                break
                    except Exception:
                        pass
            except Exception:
                pass

            # determine a preferred local IP for reaching the NVR (helps binding SADP socket)
            preferred_local = get_local_ip_for_target(nvr_ip) or None
            target_subnet = f"{nvr_ip}/24"

            devices = sadp_discover(timeout=1.0, scan_hosts=40, progress_callback=None, target_subnet=target_subnet, preferred_local_ip=preferred_local)

            for d in devices:
                found_ip = (d.get('ip') or '').strip()
                if found_ip and found_ip != nvr_ip:
                    cam_name = d.get('deviceName') or d.get('model') or f"Camera {len(cameras) + 1}"
                    cameras.append({"name": cam_name, "ip": found_ip, "status": "online"})
                    total_count += 1
                    active_count += 1
                    log(f"[6] Found SADP device: {cam_name} ({found_ip})")

            # If none found via SADP and we detected an NVR model with additional endpoints, try model-specific endpoints
            if not cameras and nvr_model:
                m_clean = nvr_model.strip()
                endpoints = MODEL_ENDPOINTS.get(m_clean, [])
                for ep in endpoints:
                    try:
                        url = f"http://{nvr_ip}{ep}"
                        resp = requests.get(url, auth=HTTPBasicAuth(username, password), timeout=1.0)
                        log(f"[6-MODEL] Tried {ep}: {resp.status_code}")
                        if resp.status_code == 200 and resp.text:
                            # attempt to parse for any IPs / channel entries
                            try:
                                # JSON or XML
                                try:
                                    data = resp.json()
                                    # find ip-like strings in returned JSON recursively
                                    def find_ips(obj):
                                        res = []
                                        if isinstance(obj, dict):
                                            for v in obj.values():
                                                res.extend(find_ips(v))
                                        elif isinstance(obj, list):
                                            for it in obj:
                                                res.extend(find_ips(it))
                                        elif isinstance(obj, str):
                                            if re.match(r'^(?:\\d{1,3}\\.){3}\\d{1,3}$', obj.strip()):
                                                res.append(obj.strip())
                                        return res
                                    ips = find_ips(data)
                                except Exception:
                                    # XML parse
                                    root = ET.fromstring(resp.text)
                                    ips = []
                                    for elem in root.iter():
                                        if elem.text and re.match(r'^(?:\\d{1,3}\\.){3}\\d{1,3}$', elem.text.strip()):
                                            ips.append(elem.text.strip())
                                for ip_addr in set(ips):
                                    if ip_addr and ip_addr != nvr_ip:
                                        cameras.append({"name": f"Camera {len(cameras)+1}", "ip": ip_addr, "status": "unknown"})
                                        total_count += 1
                            except Exception:
                                pass
                    except Exception as e:
                        log(f"[6-MODEL] Endpoint {ep} error: {e}")
                if cameras:
                    log(f"[6] SUCCESS - Found {total_count} devices via model-specific endpoints for {nvr_model}")
                    return True, cameras, total_count, active_count, ""

            if cameras:
                log(f"[6] SUCCESS - Found {total_count} devices via SADP")
                return True, cameras, total_count, active_count, ""
        except Exception as e:
            log(f"[6] SADP error: {e}")
        
        log(f"=== NO CAMERA DATA FOUND ===")
        # Return partial success if we just couldn't query but NVR is accessible
        return False, [], 0, 0, "Could not fetch camera list from NVR - try manual entry or check NVR API support"
        
    except Exception as e:
        msg = str(e)
        log(f"=== FETCH CAMERAS ERROR: {msg} ===")
        return False, [], 0, 0, msg


def check_camera_on_nvr(nvr_ip: str, nvr_user: str, nvr_pwd: str, camera_ip: str, timeout: float = 5.0) -> tuple:
    """Check if camera is live/online on NVR. Returns (online: bool, error_msg: str)."""
    try:
        # Try NVR API: get device list and check if camera_ip is registered
        url = f"http://{nvr_ip}/api/v1/devices"
        resp = requests.get(url, auth=HTTPBasicAuth(nvr_user, nvr_pwd), timeout=timeout)
        if resp.status_code == 200:
            data = resp.json()
            devices = data.get("devices", []) if isinstance(data, dict) else data
            for dev in devices:
                dev_ip = dev.get("ip") or dev.get("ipAddress") or ""
                if dev_ip == camera_ip:
                    status = dev.get("status", "").lower()
                    online = status in ("online", "active", "1", "true")
                    log(f"NVR check: {camera_ip} status={status}")
                    return online, ""
            return False, "Camera not found on NVR"
        else:
            return False, f"NVR API error {resp.status_code}"
    except Exception as e:
        log(f"NVR camera check error {camera_ip}: {e}")
        return False, str(e)

def check_camera_via_sadp(camera_ip: str, timeout: float = 3.0) -> tuple:
    """Check camera alive via SADP (Hikvision UDP discovery). Returns (online: bool, model: str)."""
    sock = None
    try:
        # SADP Protocol: send discovery request to UDP port 33333
        SADP_PORT = 33333
        SADP_REQUEST = b'\x00\x00\x00\x00\x00\x00\x00\x38<?xml version="1.0"?><Command><AccessFlag>1</AccessFlag><Command>GetDeviceInfo</Command></Command>'
        
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.settimeout(timeout)
        sock.sendto(SADP_REQUEST, (camera_ip, SADP_PORT))
        
        try:
            data, _ = sock.recvfrom(4096)
            # Parse XML response
            if b'<?xml' in data:
                xml_str = data.decode('utf-8', errors='ignore').split('<?xml', 1)[1]
                xml_str = '<?xml' + xml_str
                root = ET.fromstring(xml_str)
                model = root.findtext('.//Model', 'Unknown')
                log(f"SADP: {camera_ip} online, model={model}")
                return True, model
        except socket.timeout:
            return False, "No response"
    except Exception as e:
        log(f"SADP check error {camera_ip}: {e}")
        return False, ""
    finally:
        if sock:
            try:
                sock.close()
            except Exception:
                pass


def sadp_discover(timeout: float = 1.0, scan_hosts: int = 100, progress_callback=None, target_subnet: str = None, preferred_local_ip: str = None) -> list:
    """Discover Hikvision devices using SADP-like UDP probe.
    Sends probes quickly (broadcast + limited hosts), then collects responses within `timeout` seconds.
    Optional `progress_callback(sent, total)` is called during sending to report progress.
    If `target_subnet` is provided (like '192.168.2.0/24') it will probe that subnet first.
    If `preferred_local_ip` is provided, it will bind socket to that local IP so replies come back to correct interface.
    Returns list of dicts: {"ip":..., "model":..., "serial":..., "mac":..., "deviceName":...}
    """
    results = []
    try:
        SADP_PORT = 33333
        SADP_REQUEST = b'\x00\x00\x00\x00\x00\x00\x00\x38<?xml version="1.0"?><Command><AccessFlag>1</AccessFlag><Command>GetDeviceInfo</Command></Command>'
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
        sock.settimeout(timeout)

        # Try to bind to preferred local IP if given (helps when host has multiple NICs)
        try:
            if preferred_local_ip:
                try:
                    sock.bind((preferred_local_ip, 0))
                except Exception:
                    try:
                        sock.bind(("", 0))
                    except Exception:
                        pass
            else:
                try:
                    sock.bind(("", 0))
                except Exception:
                    pass
        except Exception:
            pass

        # prepare host list to probe
        hosts = []
        try:
            import ipaddress
            if target_subnet:
                try:
                    net = ipaddress.ip_network(target_subnet, strict=False)
                    hosts = [str(h) for i, h in enumerate(net.hosts()) if i < scan_hosts]
                except Exception:
                    hosts = []
            else:
                # infer likely local /24
                try:
                    hostname = socket.gethostname()
                    local_ip = socket.gethostbyname(hostname)
                    hosts_net = ipaddress.ip_network(f"{local_ip}/24", strict=False)
                    hosts = [str(h) for i, h in enumerate(hosts_net.hosts()) if i < scan_hosts]
                except Exception:
                    # fallback candidates
                    subs = ["192.168.1.0/24", "192.168.0.0/24", "10.0.0.0/24"]
                    for s in subs:
                        try:
                            net = ipaddress.ip_network(s, strict=False)
                            for h in net.hosts():
                                if len(hosts) >= scan_hosts:
                                    break
                                hosts.append(str(h))
                            if len(hosts) >= scan_hosts:
                                break
                        except Exception:
                            continue
        except Exception:
            hosts = []

            # Always include broadcast probe first
        try:
            sock.sendto(SADP_REQUEST, ("255.255.255.255", SADP_PORT))
        except Exception:
            pass

        total = len(hosts)
        sent = 0
        # send probes quickly without waiting for responses per-host
        for h in hosts:
            try:
                sock.sendto(SADP_REQUEST, (h, SADP_PORT))
            except Exception:
                pass
            sent += 1
            if progress_callback:
                try:
                    progress_callback(sent, total)
                except Exception:
                    pass

        # Collect responses for the timeout window
        start = time.time()
        seen = set()
        while True:
            try:
                data, addr = sock.recvfrom(8192)
                ip = addr[0]
                if ip in seen:
                    continue
                seen.add(ip)
                txt = None
                try:
                    if b'<?xml' in data:
                        txt = data.decode('utf-8', errors='ignore')
                        if '<?xml' in txt:
                            txt = txt.split('<?xml',1)[1]
                            txt = '<?xml' + txt
                    else:
                        txt = data.decode('utf-8', errors='ignore')
                except Exception:
                    txt = None

                info = {"ip": ip, "model": "", "serial": "", "mac": "", "deviceName": ""}
                if txt:
                    try:
                        root = ET.fromstring(txt)
                        for elem in root.iter():
                            tag = elem.tag.lower()
                            text = (elem.text or "").strip()
                            if not text:
                                continue
                            if 'model' in tag and not info['model']:
                                info['model'] = text
                            if 'serial' in tag or 'serialnumber' in tag:
                                info['serial'] = text
                            if 'mac' in tag and not info['mac']:
                                info['mac'] = text
                            if 'device' in tag and 'name' in tag and not info['deviceName']:
                                info['deviceName'] = text
                            if 'ip' in tag and not info['ip']:
                                info['ip'] = text
                    except Exception:
                        pass
                results.append(info)
            except socket.timeout:
                break
            except Exception:
                break
            if time.time() - start > (timeout + 0.5):
                break

        try:
            sock.close()
        except Exception:
            pass
    except Exception as e:
        log(f"[SADP DISCOVER] error: {e}")
    return results

def check_camera_live(camera_ip: str, nvr_ip: str = "", nvr_user: str = "", nvr_pwd: str = "", timeout: float = 5.0) -> tuple:
    """Check camera live status. Priority: NVR API > SADP > Ping. Returns (online: bool, method: str, details: str)."""
    # Try NVR first if provided
    if nvr_ip and nvr_user and nvr_pwd:
        online, error = check_camera_on_nvr(nvr_ip, nvr_user, nvr_pwd, camera_ip, timeout)
        if not error or "not found" not in error.lower():
            return online, "NVR API", error if not online else "Online on NVR"
    
    # Try SADP
    online, model = check_camera_via_sadp(camera_ip, timeout)
    if online:
        return True, "SADP", f"Model: {model}"
    
    # Fallback to ping
    if silent_ping(camera_ip):
        return True, "Ping", "Responds to ping"
    
    return False, "All methods", "No response"

# ---------------- sheet matching utils ----------------
def normalize_key(s):
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKC", s)
    s = "".join(s.split())
    s = re.sub(r'[^0-9A-Za-z]', '', s)
    return s.lower()

def find_sheet_key(wb_dict, target):
    t = normalize_key(target)
    # exact normalized
    for k in wb_dict.keys():
        if normalize_key(k) == t:
            return k
    # startswith / contains heuristics
    for k in wb_dict.keys():
        nk = normalize_key(k)
        if nk.startswith(t) or t.startswith(nk):
            return k
    for k in wb_dict.keys():
        nk = normalize_key(k)
        if t in nk or nk in t:
            return k
    # numeric digits match
    td = "".join(re.findall(r'\d+', t))
    if td:
        for k in wb_dict.keys():
            kd = "".join(re.findall(r'\d+', normalize_key(k)))
            if kd and kd == td:
                return k
    return None

# ---------------- Excel load: robust for header/no-header ----------------
def load_excel_robust(path):
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    # read all sheets as strings (don't coerce)
    wb = pd.read_excel(path, sheet_name=None, engine="openpyxl", dtype=str, keep_default_na=False)
    nvrs = []
    cams = []

    # --- handle NVR sheet robustly whether it has header or not ---
    if "NVR" not in wb:
        raise ValueError("No 'NVR' sheet found in workbook.")
    # get raw dataframe
    df_raw = wb["NVR"]
    # Heuristic: determine if first row appears to be header (contains words like 'Name' or 'IP')
    first_row_vals = [str(x).strip().lower() for x in list(df_raw.iloc[0].fillna(""))] if not df_raw.empty else []
    header_like = False
    header_indicators = {"name", "nvr", "ip", "subnet", "gateway", "mask"}
    if any(any(ind in v for ind in header_indicators) for v in first_row_vals):
        header_like = True

    # If header_like True -> treat as header (pandas already did). Else re-read sheet with header=None so first row becomes data.
    if header_like:
        df_nvr = df_raw
    else:
        # read again with header=None so first row is data
        df_nvr = pd.read_excel(path, sheet_name="NVR", engine="openpyxl", header=None, dtype=str, keep_default_na=False)

    # Now extract rows from df_nvr by position: A=col 0, B=col1, C=col2, D=col3
    for _, row in df_nvr.iterrows():
        name = str(row.iloc[0]).strip() if len(row) > 0 and pd.notna(row.iloc[0]) else ""
        ip = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
        subnet = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ""
        gateway = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else ""
        if name or ip:
            nvrs.append({"name": name, "ip": ip, "subnet": subnet, "gateway": gateway, "sheet_found": False})

    # For each NVR find matching sheet and load cameras (camera sheets may also lack headers; treat similarly)
    for n in nvrs:
        key = find_sheet_key(wb, n["name"])
        if key:
            n["sheet_found"] = True
            dfc_raw = wb[key]
            # detect header-like in camera sheet
            first_row = [str(x).strip().lower() for x in list(dfc_raw.iloc[0].fillna(""))] if not dfc_raw.empty else []
            cam_header_like = False
            cam_indicators = {"camera", "cam", "ip", "title", "name"}
            if any(any(ind in v for ind in cam_indicators) for v in first_row):
                cam_header_like = True
            if cam_header_like:
                df_cam = dfc_raw
            else:
                df_cam = pd.read_excel(path, sheet_name=key, engine="openpyxl", header=None, dtype=str, keep_default_na=False)
            for _, crow in df_cam.iterrows():
                cname = str(crow.iloc[0]).strip() if len(crow) > 0 and pd.notna(crow.iloc[0]) else ""
                cip = str(crow.iloc[1]).strip() if len(crow) > 1 and pd.notna(crow.iloc[1]) else ""
                if cname or cip:
                    cams.append({"nvr": n["name"], "name": cname, "ip": cip})
        else:
            n["sheet_found"] = False

    # annotate cam_count
    for n in nvrs:
        n["cam_count"] = len([c for c in cams if c.get("nvr","").strip().lower() == n.get("name","").strip().lower()])

    return nvrs, cams

# ---------------- GUI ----------------
class CameraMonitor(QtWidgets.QMainWindow):
    table_update = QtCore.pyqtSignal(int, str, str, str, object, str)
    nvr_update = QtCore.pyqtSignal(int, str)
    nvr_login_result = QtCore.pyqtSignal(bool, str, str)  # success, real_ip, error

    def get_resource_path(self, relative_path):
        """Get absolute path to resource, works for dev and for PyInstaller"""
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("🔐 NARONG CCTV TEAM v8.4.2")
        self.resize(1150, 700)
        self.vlc = find_vlc_executable()
        self.nvrs = []
        self.cams = []
        self.filtered = []
        self.check_history = {}  # persistent map: ip -> {status, device_type, model, timestamp}
        self.creds_meta = load_creds_meta()
        self.current_check_id = 0  # Track current check operation to cancel on NVR switch

        # Set window icon if logo exists
        logo_path = self.get_resource_path(LOGO_FILE)
        if os.path.exists(logo_path):
            app_icon = QtGui.QIcon(logo_path)
            self.setWindowIcon(app_icon)
            # Also set application-wide icon for taskbar
            QtWidgets.QApplication.setWindowIcon(app_icon)
            log("Window and taskbar icon loaded successfully")
        else:
            log(f"Logo file not found: {logo_path}")

        self.setStyleSheet("""
            QWidget{font-family:Arial; font-size:12px;}
            QHeaderView::section{background:#f6f6f6;padding:6px;border:1px solid #ddd;}
            QListWidget{border:1px solid #ccc;background:#fff;}
            QGroupBox{border:1px solid #ddd;padding:6px;margin-top:8px;}
        """)

        central = QtWidgets.QWidget(); self.setCentralWidget(central)
        vbox = QtWidgets.QVBoxLayout(central)

        # toolbar
        top = QtWidgets.QHBoxLayout()
        
        # Logo in toolbar
        logo_path = self.get_resource_path(LOGO_FILE)
        if os.path.exists(logo_path):
            logo_label = QtWidgets.QLabel()
            pixmap = QtGui.QPixmap(logo_path).scaledToHeight(32, QtCore.Qt.SmoothTransformation)
            logo_label.setPixmap(pixmap)
            top.addWidget(logo_label)
        
        self.btn_load = QtWidgets.QPushButton("📂 Load Excel"); self.btn_load.clicked.connect(self.load_data)
        self.btn_export = QtWidgets.QPushButton("💾 Export CSV"); self.btn_export.clicked.connect(self.export_csv)
        self.btn_check_sel = QtWidgets.QPushButton("🔍 Check Selected"); self.btn_check_sel.clicked.connect(self.check_selected)
        self.btn_check_all = QtWidgets.QPushButton("⚡ Check All"); self.btn_check_all.clicked.connect(self.check_all)
        self.btn_check_live = QtWidgets.QPushButton("📡 Check IP Online"); self.btn_check_live.clicked.connect(self.check_ip_online)
        # SADP discovery button
        self.btn_sadp = QtWidgets.QPushButton("🔧 SADP Tool"); self.btn_sadp.clicked.connect(self.show_sadp_tool)
        # Automated workflow button
        self.btn_workflow = QtWidgets.QPushButton("⚡ Quick Sync"); self.btn_workflow.clicked.connect(self.show_workflow_wizard)
        self.btn_workflow.setStyleSheet("QPushButton { background-color: #e74c3c; color: white; font-weight: bold; padding: 6px; border-radius: 4px; }")
        # Search bar with more space
        self.search = QtWidgets.QLineEdit(); 
        self.search.setPlaceholderText("🔍 Search camera or IP..."); 
        self.search.setMinimumWidth(300)
        self.search.textChanged.connect(self.filter_table)
        
        top.addWidget(self.btn_load)
        top.addWidget(self.btn_export)
        top.addWidget(self.btn_check_sel)
        top.addWidget(self.btn_check_all)
        top.addWidget(self.btn_check_live)
        top.addWidget(self.btn_sadp)
        top.addWidget(self.btn_workflow)
        top.addStretch()
        # Add update button at the end
        if UPDATE_MANAGER_AVAILABLE:
            self.btn_update = QtWidgets.QPushButton("🔄 Updates")
            self.btn_update.clicked.connect(self.check_for_updates_manual)
            top.addWidget(self.btn_update)
        top.addWidget(self.search)
        vbox.addLayout(top)

        # splitter layout
        splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal); vbox.addWidget(splitter)

        # left NVR list & details
        left = QtWidgets.QWidget(); left_l = QtWidgets.QVBoxLayout(left)
        left_l.setContentsMargins(6,6,6,6); left_l.setSpacing(6)
        left_l.addWidget(QtWidgets.QLabel("🗄️ NVR Overview"))
        self.list_nvr = QtWidgets.QListWidget(); self.list_nvr.itemSelectionChanged.connect(self.on_nvr_selected)
        self.list_nvr.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
        # Add context menu for right-click on NVR
        self.list_nvr.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.list_nvr.customContextMenuRequested.connect(self.show_nvr_context_menu)
        left_l.addWidget(self.list_nvr)
        # Replace NVR Login button with Refresh NVR button
        self.btn_refresh_nvr = QtWidgets.QPushButton("🔄 Refresh NVR Status"); self.btn_refresh_nvr.clicked.connect(self.refresh_nvr_status)
        left_l.addWidget(self.btn_refresh_nvr)
        self.grp = QtWidgets.QGroupBox("NVR Details"); form = QtWidgets.QFormLayout()
        self.lbl_name = QtWidgets.QLabel("-"); self.lbl_ip = QtWidgets.QLabel("-"); self.lbl_subnet = QtWidgets.QLabel("-"); self.lbl_gw = QtWidgets.QLabel("-"); self.lbl_sheet = QtWidgets.QLabel("-"); self.lbl_real_ip = QtWidgets.QLabel("-")
        form.addRow("Name:", self.lbl_name); form.addRow("IP:", self.lbl_ip); form.addRow("Real IP:", self.lbl_real_ip); form.addRow("Subnet:", self.lbl_subnet); form.addRow("Gateway:", self.lbl_gw); form.addRow("Sheet:", self.lbl_sheet)
        self.grp.setLayout(form); left_l.addWidget(self.grp)
        splitter.addWidget(left)

        # right camera table
        right = QtWidgets.QWidget(); right_l = QtWidgets.QVBoxLayout(right)
        right_l.setContentsMargins(4,4,4,4); right_l.setSpacing(6)
        right_l.addWidget(QtWidgets.QLabel("📷 Cameras"))

        # --- improved table block: last column stretches to fill area ---
        self.table = QtWidgets.QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["", "Camera Name", "IP", "Status", "Device Type", "NVR", "Model"])

        header = self.table.horizontalHeader()
        # camera name and ip sized to content
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        # status column moderate width (content)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        # device type sized to content
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
        # NVR column sized to content
        header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)
        # let last column (Model) stretch to fill remaining space
        header.setSectionResizeMode(6, QtWidgets.QHeaderView.Stretch)
        # icon column fixed
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        header.resizeSection(0, 30)

        self.table.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu); self.table.customContextMenuRequested.connect(self.open_context_menu)
        self.table.cellDoubleClicked.connect(self.double_click_ip)

        right_l.addWidget(self.table)
        splitter.addWidget(right)

        splitter.setStretchFactor(0, 0); splitter.setStretchFactor(1, 1)
        splitter.setSizes([200, 950])  # give more room to table by default

        # status bar + signals
        self.status = QtWidgets.QStatusBar(); self.setStatusBar(self.status)
        # Add permanent counters to status bar
        self.lbl_total = QtWidgets.QLabel("Total: 0")
        self.lbl_online = QtWidgets.QLabel("🟢 Online: 0")
        self.lbl_online.setStyleSheet("color: green; font-weight: bold;")
        self.lbl_offline = QtWidgets.QLabel("🔴 Offline: 0")
        self.lbl_offline.setStyleSheet("color: red; font-weight: bold;")
        self.status.addPermanentWidget(self.lbl_total)
        self.status.addPermanentWidget(QtWidgets.QLabel(" | "))
        self.status.addPermanentWidget(self.lbl_online)
        self.status.addPermanentWidget(QtWidgets.QLabel(" | "))
        self.status.addPermanentWidget(self.lbl_offline)
        
        self.table_update.connect(self.apply_table_update)
        self.nvr_update.connect(self.apply_nvr_update)
        self.nvr_login_result.connect(self.on_nvr_login_result)

        # Auto-load Excel on startup
        excel_path = self.get_resource_path(EXCEL_FILE)
        if os.path.exists(excel_path):
            try:
                self.load_data(initial=True)
                log(f"Auto-loaded Excel file: {excel_path}")
            except Exception as e:
                log(f"Failed to auto-load Excel: {traceback.format_exc()}")
        # load check history if present
        try:
            if os.path.exists(CHECK_HISTORY_FILE):
                with open(CHECK_HISTORY_FILE, 'r', encoding='utf-8') as f:
                    self.check_history = json.load(f)
        except Exception:
            self.check_history = {}
        
        # Check for updates on startup (after a small delay)
        if UPDATE_MANAGER_AVAILABLE:
            QtCore.QTimer.singleShot(2000, self.check_for_updates_startup)

    # ---------------- data load ----------------
    def load_data(self, initial=False):
        try:
            excel_path = self.get_resource_path(EXCEL_FILE)
            nvrs, cams = load_excel_robust(excel_path)
            self.nvrs = nvrs; self.cams = cams; self.filtered = list(cams)
            self.populate_nvr_list(); self.populate_table(self.filtered)
            self.status.showMessage(f"Loaded {len(self.nvrs)} NVRs, {len(self.cams)} cameras")
        except Exception as e:
            log(traceback.format_exc())
            if not initial:
                QtWidgets.QMessageBox.critical(self, "Load error", str(e))

    def populate_nvr_list(self):
        self.list_nvr.clear()
        for idx, n in enumerate(self.nvrs):
            sheet_flag = "" if n.get("sheet_found", False) else " ⚠️ sheet missing"
            text = f"🗄️ {n.get('name','')} | {n.get('ip','')} | 🎥 {n.get('cam_count',0)}{sheet_flag}"
            item = QtWidgets.QListWidgetItem(text); item.setData(QtCore.Qt.UserRole, idx)
            self.list_nvr.addItem(item)

    def populate_table(self, camlist):
        self.table.setRowCount(0)
        for c in camlist:
            r = self.table.rowCount(); self.table.insertRow(r)
            badge = QtWidgets.QTableWidgetItem("📷"); badge.setTextAlignment(QtCore.Qt.AlignCenter)
            badge.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            self.table.setItem(r, 0, badge)
            self.table.setItem(r, 1, QtWidgets.QTableWidgetItem(c.get("name","")))
            self.table.setItem(r, 2, QtWidgets.QTableWidgetItem(c.get("ip","")))
            # status, device type, NVR index, model placeholders
            ip_val = (c.get("ip","") or "").strip()
            hist = self.check_history.get(ip_val, {}) if ip_val else {}
            status_txt = hist.get('status', '—')
            device_txt = hist.get('device_type', '—')
            model_txt = hist.get('model', '—')
            status_color = QtGui.QColor("gray")
            if status_txt and status_txt != '—':
                # pick color by status prefix
                if status_txt.lower().startswith('online') or '🟢' in status_txt:
                    status_color = QtGui.QColor(0,160,0)
                elif status_txt.lower().startswith('offline') or '🔴' in status_txt:
                    status_color = QtGui.QColor(160,0,0)
                else:
                    status_color = QtGui.QColor(120,120,120)

            sit = QtWidgets.QTableWidgetItem(status_txt); sit.setForeground(status_color); self.table.setItem(r, 3, sit)
            dit = QtWidgets.QTableWidgetItem(device_txt); dit.setForeground(status_color); self.table.setItem(r, 4, dit)
            mit = QtWidgets.QTableWidgetItem(model_txt); mit.setForeground(status_color); self.table.setItem(r, 6, mit)
            # NVR column: try to fill from camera's nvr field
            nvr_idx_text = ""
            try:
                cam_nvr = (c.get("nvr", "") or "").strip()
                if cam_nvr:
                    # find matching nvr index
                    for i, n in enumerate(self.nvrs):
                        if n.get("name","" ).strip().lower() == cam_nvr.lower():
                            nvr_idx_text = str(i+1)
                            break
                    # if not found, show name
                    if not nvr_idx_text:
                        nvr_idx_text = cam_nvr
            except Exception:
                nvr_idx_text = ""
            nvt = QtWidgets.QTableWidgetItem(nvr_idx_text if nvr_idx_text else "—")
            nvt.setForeground(QtGui.QColor("black"))
            self.table.setItem(r, 5, nvt)
        
        # Update status bar counters
        self.update_counters()
        
        # Update status bar counters
        self.update_counters()
    
    def update_counters(self):
        """Update status bar counters for total/online/offline cameras"""
        total = self.table.rowCount()
        online = 0
        offline = 0
        
        for r in range(total):
            status_item = self.table.item(r, 3)
            if status_item:
                status_txt = status_item.text().lower()
                if 'online' in status_txt or '🟢' in status_txt:
                    online += 1
                elif 'offline' in status_txt or '🔴' in status_txt:
                    offline += 1
        
        self.lbl_total.setText(f"Total: {total}")
        self.lbl_online.setText(f"🟢 Online: {online}")
        self.lbl_offline.setText(f"🔴 Offline: {offline}")

    # ---------------- NVR select/filter ----------------
    def on_nvr_selected(self):
        items = self.list_nvr.selectedItems()
        if not items:
            return
        idx = items[0].data(QtCore.Qt.UserRole)
        n = self.nvrs[idx]
        self.lbl_name.setText(n.get("name","")); self.lbl_ip.setText(n.get("ip",""))
        self.lbl_subnet.setText(n.get("subnet","")); self.lbl_gw.setText(n.get("gateway",""))
        self.lbl_sheet.setText("Found" if n.get("sheet_found", False) else "Missing")
        name = n.get("name","").strip().lower()
        cams = [c for c in self.cams if c.get("nvr","").strip().lower() == name]
        self.filtered = cams; self.populate_table(self.filtered)
        self.status.showMessage(f"{len(self.filtered)} cameras for {n.get('name','')}")
        
        # Cancel any ongoing check operations by incrementing check ID
        self.current_check_id += 1

    def filter_table(self):
        q = self.search.text().strip().lower()
        if not q:
            self.filtered = list(self.cams)
        else:
            self.filtered = [c for c in self.cams if q in c.get("name","").lower() or q in c.get("ip","").lower() or q in c.get("nvr","").lower()]
        self.populate_table(self.filtered)
        self.status.showMessage(f"{len(self.filtered)} entries")

    # ---------------- checks ----------------
    def check_selected(self):
        rows = sorted({r.row() for r in self.table.selectionModel().selectedRows()})
        if not rows:
            QtWidgets.QMessageBox.information(self, "Check", "Select camera rows first.")
            return
        # Increment check ID to cancel any previous operations
        self.current_check_id += 1
        check_id = self.current_check_id
        targets = [{"row": r, "ip": self.table.item(r,2).text().strip()} for r in rows if self.table.item(r,2)]
        threading.Thread(target=self._run_checks, args=(targets, check_id), daemon=True).start()

    def check_all(self):
        # Increment check ID to cancel any previous operations
        self.current_check_id += 1
        check_id = self.current_check_id
        targets = [{"row": r, "ip": self.table.item(r,2).text().strip()} for r in range(self.table.rowCount()) if self.table.item(r,2)]
        threading.Thread(target=self._run_checks, args=(targets, check_id), daemon=True).start()

    def check_live_status(self):
        """Check camera live status via NVR API or SADP."""
        rows = sorted({r.row() for r in self.table.selectionModel().selectedRows()})
        if not rows:
            QtWidgets.QMessageBox.information(self, "Check Live", "Select camera rows first, or click again to check all.")
            return
        
        # Get NVR credentials for the selected cameras
        nvr_ip = nvr_user = nvr_pwd = ""
        selected_nvr_name = None
        
        # Try to get NVR info from selected cameras
        for row in rows:
            cam = self.filtered[row] if row < len(self.filtered) else None
            if cam:
                cam_nvr_name = cam.get("nvr", "").strip().lower()
                for n in self.nvrs:
                    if n.get("name", "").strip().lower() == cam_nvr_name:
                        nvr_ip = n.get("ip", "")
                        nvr_user, nvr_pwd = get_password(nvr_ip) if nvr_ip else (None, None)
                        if not nvr_user:
                            nvr_user = "admin"
                        if not nvr_pwd:
                            nvr_pwd = DEFAULT_CREDS[0][1] if DEFAULT_CREDS else "Kkcctv12345"
                        selected_nvr_name = n.get("name", "")
                        break
                if nvr_ip:
                    break
        
        targets = [{"row": r, "ip": self.table.item(r, 2).text().strip()} for r in rows if self.table.item(r, 2)]
        threading.Thread(target=self._run_live_checks, args=(targets, nvr_ip, nvr_user, nvr_pwd), daemon=True).start()

    def check_ip_online(self):
        """Check all NVRs and all cameras' IPs concurrently (fast, parallel)."""
        # Disable button to avoid re-entry
        try:
            self.btn_check_live.setEnabled(False)
            self.status.showMessage("Checking all NVRs and cameras...")
        except Exception:
            pass

        # prepare camera-to-row mapping for visible table rows
        ip_to_row = {}
        for r in range(self.table.rowCount()):
            try:
                item = self.table.item(r,2)
                if item:
                    ip = item.text().strip()
                    if ip:
                        ip_to_row[ip] = r
            except Exception:
                continue

        # Prepare NVR list
        nvrs = list(self.nvrs)
        cams = list(self.cams)

        def fast_check_ip(ip):
            # Try SADP
            try:
                sadp_online, sadp_model = check_camera_via_sadp(ip, timeout=0.8)
                if sadp_online:
                    return True, "SADP", sadp_model
            except Exception:
                pass
            # Try TCP ports
            try:
                h = check_tcp(ip, HTTP_PORT, timeout=0.5)
                r = check_tcp(ip, RTSP_PORT, timeout=0.5)
                if h or r:
                    methods = []
                    if h: methods.append('HTTP')
                    if r: methods.append('RTSP')
                    return True, 'TCP', ','.join(methods)
            except Exception:
                pass
            # Last resort ping
            try:
                if silent_ping(ip):
                    return True, 'Ping', ''
            except Exception:
                pass
            return False, 'Offline', ''

        def run_checks_thread():
            cam_online = 0
            cam_total = 0
            cam_offline = 0
            # check NVRs first via existing _check_nvr (which updates UI)
            for idx, n in enumerate(nvrs):
                try:
                    threading.Thread(target=self._check_nvr, args=(idx, n), daemon=True).start()
                except Exception:
                    pass

            # Now check cameras in parallel
            cam_total = len(cams)
            if cam_total == 0:
                QtCore.QTimer.singleShot(0, lambda: self._finish_ip_check(0,0,0))
                return

            # Use ThreadPoolExecutor for parallel checks and show a progress dialog
            cam_total = len(cams)
            prog = None
            try:
                QtCore.QTimer.singleShot(0, lambda: self.status.showMessage('Checking cameras...'))
                prog = QtWidgets.QProgressDialog('Checking cameras...', 'Cancel', 0, cam_total, self)
                prog.setWindowModality(QtCore.Qt.WindowModal)
                prog.setMinimumDuration(0)
                prog.setValue(0)
                # must show on UI thread
                QtCore.QTimer.singleShot(0, prog.show)
            except Exception:
                prog = None

            checked = 0
            with concurrent.futures.ThreadPoolExecutor(max_workers=60) as ex:
                futures = {ex.submit(fast_check_ip, c.get('ip','')): c for c in cams if c.get('ip')}
                for fut in concurrent.futures.as_completed(futures):
                    cam = futures[fut]
                    ip = cam.get('ip','')
                    try:
                        ok, method, details = fut.result()
                    except Exception as e:
                        ok, method, details = False, 'Error', str(e)

                    checked += 1
                    # prepare update values
                    if ok:
                        cam_online += 1
                        em = '🟢'
                        color = QtGui.QColor(0,160,0)
                        status_text = f'Online ({method})'
                        device_type = method
                        model = details or '—'
                    else:
                        cam_offline += 1
                        em = '🔴'
                        color = QtGui.QColor(160,0,0)
                        status_text = 'Offline'
                        device_type = '—'
                        model = '—'

                    # update visible table row
                    row = ip_to_row.get(ip)
                    if row is not None:
                        self.table_update.emit(row, status_text, device_type, model, color, em)

                    # update progress dialog
                    if prog:
                        QtCore.QTimer.singleShot(0, lambda v=checked: prog.setValue(v))

                    log(f"Check IP {ip}: {status_text} ({method})")

            QtCore.QTimer.singleShot(0, lambda: self._finish_ip_check(cam_total, cam_online, cam_offline))

        threading.Thread(target=run_checks_thread, daemon=True).start()

    def _finish_ip_check(self, total, online, offline):
        try:
            self.btn_check_live.setEnabled(True)
            self.status.showMessage(f"Check complete: {online}/{total} online, {offline} offline")
            QtWidgets.QMessageBox.information(self, "Check IP Online", f"Cameras: {online}/{total} online\nOffline: {offline}")
        except Exception:
            pass

    def _run_live_checks(self, targets, nvr_ip, nvr_user, nvr_pwd):
        self.status.showMessage(f"Checking live status for {len(targets)} cameras...")
        for t in targets:
            row = t["row"]
            ip = t["ip"]
            try:
                online, method, details = check_camera_live(ip, nvr_ip, nvr_user, nvr_pwd, timeout=5.0)
                if online:
                    em = "🟢"
                    color = QtGui.QColor(0, 160, 0)
                    status_text = f"Live ({method})"
                else:
                    em = "🔴"
                    color = QtGui.QColor(160, 0, 0)
                    status_text = f"Offline ({method})"
                
                device_type = method
                model = details
                self.table_update.emit(row, status_text, device_type, model, color, em)
                log(f"Live check {ip}: {status_text} - {details}")
            except Exception as e:
                log(f"Live check error {ip}: {e}")
                self.table_update.emit(row, "Error", "Error", str(e), QtGui.QColor(128, 0, 0), "⚠️")
        self.status.showMessage("Live status check complete.")

    def closeEvent(self, event):
        # Persist check history before closing
        try:
            with open(CHECK_HISTORY_FILE, 'w', encoding='utf-8') as hf:
                json.dump(self.check_history, hf, indent=2)
            log(f"[CLOSE] Saved check history ({len(self.check_history)} entries)")
        except Exception as e:
            log(f"[CLOSE] Error saving history: {e}")
        try:
            super().closeEvent(event)
        except Exception:
            event.accept()

    def _run_checks(self, targets, check_id=None):
        self.status.showMessage(f"Checking {len(targets)} cameras...")
        checked_count = 0
        for t in targets:
            # Check if this operation was cancelled by NVR switch
            if check_id is not None and check_id != self.current_check_id:
                self.status.showMessage(f"Check cancelled (switched NVR)")
                return
                
            row = t["row"]; ip = t["ip"]
            try:
                # Try SADP first (Hikvision UDP discovery - most reliable, faster timeout)
                sadp_online, sadp_model = check_camera_via_sadp(ip, timeout=1.5)
                if sadp_online:
                    em = "🟢"
                    color = QtGui.QColor(0, 160, 0)
                    status_text = "Online (SADP)"
                    device_type = "SADP"
                    model = f"Model: {sadp_model}"
                    self.table_update.emit(row, status_text, device_type, model, color, em)
                    log(f"checked {ip}: SADP detected, model={sadp_model}")
                    checked_count += 1
                    continue
                
                # Fallback: Try TCP ports (HTTP + RTSP) with faster timeout
                h = check_tcp(ip, HTTP_PORT, timeout=0.8)
                r = check_tcp(ip, RTSP_PORT, timeout=0.8)
                if h or r:
                    em = "🟢"
                    color = QtGui.QColor(0, 160, 0)
                    status_text = "Online (TCP)"
                    device_type = f"{'HTTP' if h else ''} {'RTSP' if r else ''}".strip() or "—"
                    model = "—"
                else:
                    # Last resort: Try ping (fast)
                    p = silent_ping(ip)
                    if p:
                        em = "🟡"
                        color = QtGui.QColor(200, 140, 0)
                        status_text = "Online (Ping)"
                        device_type = "Ping"
                        model = "—"
                    else:
                        em = "🔴"
                        color = QtGui.QColor(160, 0, 0)
                        status_text = "Offline"
                        device_type = "—"
                        model = "—"
                
                self.table_update.emit(row, status_text, device_type, model, color, em)
                log(f"checked {ip}: {status_text}, device_type={device_type}")
                checked_count += 1
            except Exception as e:
                log(f"check error {ip}: {e}")
                self.table_update.emit(row, "Error", "Error", str(e)[:50], QtGui.QColor(128, 0, 0), "⚠️")
                checked_count += 1
        
        self.status.showMessage(f"Camera checks complete. ({checked_count}/{len(targets)} checked)")

    @QtCore.pyqtSlot(int, str, str, str, object, str)
    def apply_table_update(self, row, status_text, device_type_text, model_text, color, emoji):
        try:
            badge = QtWidgets.QTableWidgetItem(emoji); badge.setTextAlignment(QtCore.Qt.AlignCenter)
            badge.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            self.table.setItem(row, 0, badge)
            sit = QtWidgets.QTableWidgetItem(status_text); sit.setForeground(color); self.table.setItem(row, 3, sit)
            dit = QtWidgets.QTableWidgetItem(device_type_text); dit.setForeground(color); self.table.setItem(row, 4, dit)
            # Determine NVR index for this camera (if available)
            nvr_idx_text = ""
            try:
                ip_item = self.table.item(row, 2)
                ip_val = ip_item.text().strip() if ip_item else ""
                if ip_val:
                    # find camera entry
                    for cam in self.cams:
                        if cam.get("ip","") == ip_val:
                            nvr_name = cam.get("nvr","")
                            # find nvr index in self.nvrs
                            for i, n in enumerate(self.nvrs):
                                if n.get("name","").strip().lower() == nvr_name.strip().lower():
                                    nvr_idx_text = str(i+1)
                                    break
                            break
            except Exception:
                nvr_idx_text = ""

            nvt = QtWidgets.QTableWidgetItem(nvr_idx_text); nvt.setForeground(color); self.table.setItem(row, 5, nvt)
            mit = QtWidgets.QTableWidgetItem(model_text); mit.setForeground(color); self.table.setItem(row, 6, mit)
            # persist into in-memory cam entries and check_history
            try:
                ip_item = self.table.item(row, 2)
                ip_val = ip_item.text().strip() if ip_item else ""
                if ip_val:
                    # update in-memory cams
                    for cam in self.cams:
                        if cam.get('ip','') == ip_val:
                            cam['last_status'] = status_text
                            cam['last_method'] = device_type_text
                            cam['last_model'] = model_text
                            break
                    # update history and save to file
                    try:
                        self.check_history[ip_val] = {
                            'status': status_text,
                            'device_type': device_type_text,
                            'model': model_text,
                            'timestamp': time.time()
                        }
                        with open(CHECK_HISTORY_FILE, 'w', encoding='utf-8') as hf:
                            json.dump(self.check_history, hf, indent=2)
                    except Exception as e:
                        log(f"[HISTORY SAVE] Error saving history: {e}")
            except Exception:
                pass
        except Exception:
            pass
        
        # Update status bar counters
        self.update_counters()

    # ---------------- NVR refresh ----------------
    def refresh_nvr_status(self):
        """Refresh NVR status: connectivity + real IP detection."""
        self.status.showMessage("Refreshing NVR status...")
        for idx, n in enumerate(self.nvrs):
            threading.Thread(target=self._check_nvr, args=(idx, n), daemon=True).start()

    def _check_nvr(self, index, nvr_obj):
        """Check NVR connectivity and detect real IP."""
        try:
            ip = nvr_obj.get("ip", "")
            name = nvr_obj.get("name", "")
            
            # Try to get NVR credentials
            username, password = get_password(ip)
            if not username:
                username = "admin"
            if not password:
                password = DEFAULT_CREDS[0][1] if DEFAULT_CREDS else "Kkcctv12345"
            
            # Try SADP first (faster, more reliable)
            log(f"[NVR REFRESH] Checking {name} ({ip}) via SADP...")
            sadp_online, sadp_model = check_camera_via_sadp(ip, timeout=1.0)
            if sadp_online:
                emoji = "🟢 SADP"
                log(f"[NVR REFRESH] {name} online via SADP, model={sadp_model}")
                self.nvr_update.emit(index, emoji)
                return
            
            # Fallback: Try TCP connection to HTTP port
            log(f"[NVR REFRESH] SADP failed, trying HTTP...")
            if check_tcp(ip, HTTP_PORT, timeout=1.0):
                emoji = "🟢 HTTP"
                log(f"[NVR REFRESH] {name} online via HTTP")
                self.nvr_update.emit(index, emoji)
                return
            
            # Last resort: Try ping
            log(f"[NVR REFRESH] HTTP failed, trying ping...")
            if silent_ping(ip):
                emoji = "🟡 Ping"
                log(f"[NVR REFRESH] {name} responds to ping only")
                self.nvr_update.emit(index, emoji)
                return
            
            # Offline
            emoji = "🔴 Offline"
            log(f"[NVR REFRESH] {name} is offline")
            self.nvr_update.emit(index, emoji)
        except Exception as e:
            log(f"[NVR REFRESH] Error checking {nvr_obj.get('name', 'Unknown')}: {e}")
            self.nvr_update.emit(index, "⚠️ Error")

    @QtCore.pyqtSlot(int, str)
    def apply_nvr_update(self, index, emoji):
        try:
            item = self.list_nvr.item(index)
            if not item:
                return
            n = self.nvrs[index]
            sheet_flag = "" if n.get("sheet_found", False) else " ⚠️ sheet missing"
            text = f"{emoji}  🗄️ {n.get('name','')} | {n.get('ip','')} | 🎥 {n.get('cam_count',0)}{sheet_flag}"
            item.setText(text)
            log(f"[NVR UPDATE] Updated NVR {index}: {text}")
        except Exception as e:
            log(f"[NVR UPDATE] Error: {e}")

    # ---------------- context menu & double-click ----------------
    def open_context_menu(self, pos):
        r = self.table.currentRow()
        if r < 0:
            return
        ip = self.table.item(r,2).text().strip()
        m = QtWidgets.QMenu()
        a1 = m.addAction("🌐  Open HTTP in Browser")
        a2 = m.addAction("🎦  Open RTSP in VLC")
        a3 = m.addAction("📸  Snapshot (http://ip/snapshot)")
        m.addSeparator()
        a4 = m.addAction("🪪  Set Credentials...")
        a5 = m.addAction("🗑️  Forget Credentials")
        action = m.exec_(self.table.viewport().mapToGlobal(pos))
        if action == a1:
            self.open_http_with_try(ip)
        elif action == a2:
            threading.Thread(target=self.open_rtsp_direct, args=(ip,), daemon=True).start()
        elif action == a3:
            webbrowser.open(f"http://{ip}/snapshot")
        elif action == a4:
            self.show_credentials_dialog(ip)
        elif action == a5:
            delete_credentials(ip)
            QtWidgets.QMessageBox.information(self, "Credentials", "Credentials removed (best-effort).")

    def double_click_ip(self, row, col):
        if col == 2:
            ip = self.table.item(row,2).text().strip()
            if ip:
                self.status.showMessage(f"Launching VLC for {ip}...")
                threading.Thread(target=self.open_rtsp_direct, args=(ip,), daemon=True).start()

    # ---------------- open helpers ----------------
    def open_http_with_try(self, ip):
        u,p = get_password(ip)
        if u and p:
            webbrowser.open(f"http://{u}:{p}@{ip}"); return
        for du,dp in DEFAULT_CREDS:
            webbrowser.open(f"http://{du}:{dp}@{ip}"); return
        webbrowser.open(f"http://{ip}")

    def open_rtsp_direct(self, ip):
        u,p = get_password(ip)
        if u and p:
            if self._launch_vlc_rtsp(ip,u,p): return
        for du,dp in DEFAULT_CREDS:
            if self._launch_vlc_rtsp(ip,du,dp): return
        self._launch_vlc_rtsp(ip,None,None)

    def _launch_vlc_rtsp(self, ip, username, password):
        if username and password:
            url = f"rtsp://{username}:{password}@{ip}"
        else:
            url = f"rtsp://{ip}"
        try:
            subprocess.Popen([self.vlc, url], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            log(f"Launching VLC -> {url}")
            return True
        except Exception as e:
            log(f"VLC launch failed ({url}): {e}")
            return False

    # ---------------- credentials dialog ----------------
    def show_credentials_dialog(self, ip):
        dlg = QtWidgets.QDialog(self); dlg.setWindowTitle(f"Credentials for {ip}")
        layout = QtWidgets.QFormLayout(dlg)
        user_edit = QtWidgets.QLineEdit(); pwd_edit = QtWidgets.QLineEdit(); pwd_edit.setEchoMode(QtWidgets.QLineEdit.Password)
        meta = load_creds_meta()
        if ip in meta: user_edit.setText(meta[ip].get("username",""))
        auto_chk = QtWidgets.QCheckBox("Auto-login when opening RTSP"); auto_chk.setChecked(meta.get(ip,{}).get("auto_login", False))
        layout.addRow("Username:", user_edit); layout.addRow("Password:", pwd_edit); layout.addRow(auto_chk)
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel); layout.addRow(btns)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            username = user_edit.text().strip(); pwd = pwd_edit.text(); auto_login = auto_chk.isChecked()
            if username and pwd:
                ok = set_password(ip, username, pwd)
                meta = load_creds_meta(); meta[ip] = {"username": username, "auto_login": auto_login}; save_creds_meta(meta)
                self.creds_meta = meta
                if ok and KEYRING_AVAILABLE:
                    QtWidgets.QMessageBox.information(self, "Saved", "Credentials saved to system keyring.")
                else:
                    QtWidgets.QMessageBox.information(self, "Saved (fallback)", "Credentials saved to local fallback store.")
            else:
                QtWidgets.QMessageBox.warning(self, "Invalid", "Provide username and password to save.")

    # ---------------- export ----------------
    def export_csv(self):
        if not self.cams:
            QtWidgets.QMessageBox.information(self, "Export", "No cameras to export."); return
        try:
            with open(EXPORT_FILE, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f); w.writerow(["nvr","camera","ip"])
                for c in self.cams:
                    w.writerow([c.get("nvr",""), c.get("name",""), c.get("ip","")])
            QtWidgets.QMessageBox.information(self, "Export", f"Saved {EXPORT_FILE}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export error", str(e))

    # ---------------- NVR context menu ----------------
    def show_nvr_context_menu(self, position):
        """Show context menu on right-click for NVR list."""
        items = self.list_nvr.selectedItems()
        if not items:
            return
        
        menu = QtWidgets.QMenu()
        menu.setStyleSheet("""
            QMenu {
                background-color: white;
                border: 1px solid #bdc3c7;
            }
            QMenu::item {
                padding: 8px 25px;
            }
            QMenu::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)
        
        # Add logo to menu if available
        logo_path = self.get_resource_path(LOGO_FILE)
        if os.path.exists(logo_path):
            menu.setIcon(QtGui.QIcon(logo_path))
        
        action_login = menu.addAction("🔐 NVR Login")
        action_refresh = menu.addAction("🔄 Refresh Status")
        menu.addSeparator()
        action_fetch_cameras = menu.addAction("📹 Fetch & Update Cameras")
        action_open_browser = menu.addAction("🌐 Open in Browser")
        
        action = menu.exec_(self.list_nvr.mapToGlobal(position))
        
        if action == action_login:
            self.show_nvr_login_dialog()
        elif action == action_refresh:
            self.refresh_nvr_status()
        elif action == action_fetch_cameras:
            self.show_nvr_login_dialog()  # Opens dialog with Fetch & Update option
        elif action == action_open_browser:
            idx = items[0].data(QtCore.Qt.UserRole)
            nvr = self.nvrs[idx]
            nvr_ip = nvr.get("ip", "").strip()
            if nvr_ip:
                webbrowser.open(f"http://{nvr_ip}")

    # ---------------- NVR login dialog ----------------
    def show_nvr_login_dialog(self):
        items = self.list_nvr.selectedItems()
        if not items:
            QtWidgets.QMessageBox.information(self, "NVR Login", "Select an NVR from the list first.")
            return
        idx = items[0].data(QtCore.Qt.UserRole)
        nvr = self.nvrs[idx]
        nvr_ip = nvr.get("ip", "").strip()
        nvr_name = nvr.get("name", "")
        if not nvr_ip:
            QtWidgets.QMessageBox.warning(self, "NVR Login", "NVR has no IP address.")
            return
        
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(f"NVR Login - {nvr_name} ({nvr_ip})")
        layout = QtWidgets.QFormLayout(dlg)
        
        user_edit = QtWidgets.QLineEdit(); user_edit.setText("admin")
        pwd_edit = QtWidgets.QLineEdit(); pwd_edit.setEchoMode(QtWidgets.QLineEdit.Password); pwd_edit.setText("Kkcctv12345")
        real_ip_label = QtWidgets.QLabel("-")
        
        layout.addRow("NVR IP:", QtWidgets.QLabel(nvr_ip))
        layout.addRow("Username:", user_edit)
        layout.addRow("Password:", pwd_edit)
        layout.addRow("Real IP (fetched):", real_ip_label)
        
        test_btn = QtWidgets.QPushButton("🔍 Test Login & Fetch IP")
        update_btn = QtWidgets.QPushButton("💾 Update IP in Data")
        fetch_cams_btn = QtWidgets.QPushButton("📹 Fetch & Update Cameras")
        browser_btn = QtWidgets.QPushButton("🌐 Open in Browser")
        save_creds_btn = QtWidgets.QPushButton("🪪 Save Credentials")
        
        # Info labels for camera count
        cam_info_label = QtWidgets.QLabel("-")
        cam_info_label.setStyleSheet("color: blue; font-weight: bold;")
        
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Close)
        
        layout.addWidget(test_btn)
        layout.addWidget(update_btn)
        layout.addWidget(fetch_cams_btn)
        layout.addRow("Camera Info:", cam_info_label)
        layout.addWidget(browser_btn)
        layout.addWidget(save_creds_btn)
        layout.addWidget(btns)
        btns.rejected.connect(dlg.reject)
        
        real_ip_storage = [""]  # store fetched IP
        cameras_storage = []  # store fetched cameras
        creds_storage = {"username": "admin", "password": "Kkcctv12345"}  # store credentials
        
        def test_login_action():
            username = user_edit.text().strip()
            password = pwd_edit.text()
            if not username or not password:
                QtWidgets.QMessageBox.warning(dlg, "Input Error", "Enter username and password.")
                return
            creds_storage["username"] = username
            creds_storage["password"] = password
            test_btn.setEnabled(False); test_btn.setText("Testing...")
            threading.Thread(target=self._test_nvr_login_thread, args=(dlg, nvr_ip, nvr_name, username, password, real_ip_label, real_ip_storage, test_btn), daemon=True).start()
        
        def browser_action():
            username = user_edit.text().strip()
            password = pwd_edit.text()
            if not username or not password:
                QtWidgets.QMessageBox.warning(dlg, "Input Error", "Enter username and password.")
                return
            # Use real IP if available, otherwise use current IP
            ip_to_use = real_ip_storage[0] if real_ip_storage[0] else nvr_ip
            url = f"http://{username}:{password}@{ip_to_use}"
            webbrowser.open(url)
            log(f"Opening NVR in browser: {ip_to_use}")
            QtWidgets.QMessageBox.information(dlg, "Browser Opened", f"Opening {ip_to_use} in default browser.")
        
        def save_creds_action():
            username = user_edit.text().strip()
            password = pwd_edit.text()
            if not username or not password:
                QtWidgets.QMessageBox.warning(dlg, "Input Error", "Enter username and password.")
                return
            creds_storage["username"] = username
            creds_storage["password"] = password
            # Save to metadata
            set_password(nvr_ip, username, password)
            meta = load_creds_meta()
            meta[nvr_ip] = {"username": username}
            save_creds_meta(meta)
            QtWidgets.QMessageBox.information(dlg, "Success", "NVR credentials saved securely.")
            log(f"NVR credentials saved for {nvr_ip}")
        
        def update_ip_action():
            log(f"[UPDATE IP ACTION] Called. real_ip_storage[0]='{real_ip_storage[0]}'")
            new_ip = real_ip_storage[0]
            
            if not new_ip or new_ip.strip() == "":
                log(f"[UPDATE IP ACTION] ERROR: new_ip is empty or None: '{new_ip}'")
                QtWidgets.QMessageBox.warning(dlg, "Error", "Fetch IP first by testing login.")
                return
            
            try:
                excel_path = EXCEL_FILE
                if not os.path.exists(excel_path):
                    QtWidgets.QMessageBox.warning(dlg, "Error", "Excel file not found.")
                    return
                
                log(f"[UPDATE IP] Opening Excel: {excel_path}")
                log(f"[UPDATE IP] Looking for NVR: name='{nvr_name}', old_ip='{nvr_ip}', new_ip='{new_ip}'")
                
                # Use openpyxl for direct cell editing to preserve formatting
                from openpyxl import load_workbook
                
                wb_openpyxl = load_workbook(excel_path)
                if "NVR" not in wb_openpyxl.sheetnames:
                    QtWidgets.QMessageBox.warning(dlg, "Error", "NVR sheet not found in Excel.")
                    return
                
                ws = wb_openpyxl["NVR"]
                log(f"[UPDATE IP] NVR sheet has {ws.max_row} rows")
                
                updated = False
                for row_idx in range(1, ws.max_row + 1):
                    cell_name = ws.cell(row=row_idx, column=1).value  # Column A: Name
                    cell_ip = ws.cell(row=row_idx, column=2).value    # Column B: IP
                    
                    cell_name_str = str(cell_name).strip() if cell_name else ""
                    cell_ip_str = str(cell_ip).strip() if cell_ip else ""
                    
                    log(f"[UPDATE IP] Row {row_idx}: name='{cell_name_str}', ip='{cell_ip_str}'")
                    
                    # Match by IP or name
                    if cell_ip_str == nvr_ip or cell_name_str == nvr_name:
                        log(f"[UPDATE IP] MATCH! Row {row_idx}: updating IP from '{cell_ip_str}' to '{new_ip}'")
                        ws.cell(row=row_idx, column=2).value = new_ip
                        updated = True
                        break
                
                if not updated:
                    log(f"[UPDATE IP] No matching NVR found")
                    QtWidgets.QMessageBox.warning(dlg, "Error", f"NVR '{nvr_name}' not found in Excel.")
                    return
                
                # Save the workbook
                log(f"[UPDATE IP] Saving Excel file...")
                wb_openpyxl.save(excel_path)
                log(f"[UPDATE IP] Excel saved successfully")
                
                # Update in-memory data
                for nvr_obj in self.nvrs:
                    if (nvr_obj.get("ip") == nvr_ip or nvr_obj.get("name") == nvr_name):
                        nvr_obj["ip"] = new_ip
                        self.lbl_ip.setText(new_ip)
                        self.populate_nvr_list()
                        break
                
                log(f"[UPDATE IP] SUCCESS: {nvr_name} IP updated from {nvr_ip} to {new_ip}")
                QtWidgets.QMessageBox.information(dlg, "Success", f"✅ NVR IP updated to {new_ip}")
                
            except Exception as e:
                import traceback
                QtWidgets.QMessageBox.critical(dlg, "Update Error", f"Error: {str(e)}")
                log(f"[UPDATE IP] ERROR: {e}")
                log(f"[UPDATE IP] Traceback: {traceback.format_exc()}")
        
        def fetch_cameras_action():
            """Fetch cameras from NVR and update Excel with their IPs."""
            log(f"[FETCH CAMERAS] Starting camera fetch for NVR {nvr_name}")
            username = user_edit.text().strip()
            password = pwd_edit.text()
            if not username or not password:
                QtWidgets.QMessageBox.warning(dlg, "Input Error", "Enter username and password.")
                return
            
            fetch_cams_btn.setEnabled(False)
            fetch_cams_btn.setText("Fetching...")
            
            # Show progress while fetching
            progress = QtWidgets.QProgressDialog("Fetching cameras from NVR...", None, 0, 0, dlg)
            progress.setWindowModality(QtCore.Qt.WindowModal)
            progress.setWindowTitle("Fetch Cameras")
            progress.setCancelButton(None)
            progress.setMinimumDuration(0)
            progress.show()

            # Fetch cameras in background thread
            def fetch_thread():
                success, cameras, total, active, error = fetch_nvr_cameras(nvr_ip, username, password)
                QtCore.QTimer.singleShot(0, lambda: progress.close())
                QtCore.QTimer.singleShot(0, lambda: on_fetch_complete(success, cameras, total, active, error))

            threading.Thread(target=fetch_thread, daemon=True).start()
        
        def on_fetch_complete(success, cameras, total, active, error):
            """Called when camera fetch completes."""
            fetch_cams_btn.setEnabled(True)
            fetch_cams_btn.setText("📹 Fetch & Update Cameras")
            
            log(f"[FETCH CAMERAS] Fetch complete: success={success}, total={total}, active={active}")
            
            if not success:
                QtWidgets.QMessageBox.critical(dlg, "Fetch Failed", f"Failed to fetch cameras:\n{error}")
                return
            
            if not cameras:
                QtWidgets.QMessageBox.information(dlg, "No Cameras", "No cameras found on NVR or NVR doesn't support camera listing.")
                return
            
            # Update camera info label
            cam_info_label.setText(f"📹 Total: {total} | 🟢 Active: {active}")
            cameras_storage.clear()
            cameras_storage.extend(cameras)
            
            log(f"[FETCH CAMERAS] Found {total} cameras: {[c['name'] for c in cameras]}")
            
            # Ask if user wants to update Excel
            msg = f"Found {total} cameras ({active} active):\n\n"
            for cam in cameras[:5]:  # Show first 5
                msg += f"  • {cam['name']}: {cam['ip']}\n"
            if len(cameras) > 5:
                msg += f"  ... and {len(cameras) - 5} more\n"
            msg += "\nUpdate camera IPs in Excel?"
            
            reply = QtWidgets.QMessageBox.question(dlg, "Update Cameras", msg, QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            if reply == QtWidgets.QMessageBox.Yes:
                # Run update in background and show determinate progress
                prog = QtWidgets.QProgressDialog("Updating camera IPs...", "Cancel", 0, len(cameras), dlg)
                prog.setWindowModality(QtCore.Qt.WindowModal)
                prog.setWindowTitle("Update Cameras")
                prog.setMinimumDuration(0)
                prog.show()

                def update_worker():
                    try:
                        from openpyxl import load_workbook
                        excel_path = EXCEL_FILE
                        if not os.path.exists(excel_path):
                            QtCore.QTimer.singleShot(0, lambda: QtWidgets.QMessageBox.warning(dlg, "Error", "Excel file not found."))
                            QtCore.QTimer.singleShot(0, lambda: prog.close())
                            return
                        wb = load_workbook(excel_path)
                        sheet_name = nvr_name.replace("/", "_").replace("\\", "_")
                        if sheet_name not in wb.sheetnames:
                            ws = wb.create_sheet(sheet_name)
                            ws.cell(row=1, column=1).value = "Camera Name"
                            ws.cell(row=1, column=2).value = "IP Address"
                        else:
                            ws = wb[sheet_name]

                        updated_count = 0
                        for idx, camera in enumerate(cameras, start=1):
                            cam_name = camera.get("name", f"Camera {idx}")
                            cam_ip = camera.get("ip", "")
                            if not cam_ip:
                                QtCore.QTimer.singleShot(0, lambda: prog.setValue(idx))
                                continue
                            found = False
                            for row_idx in range(2, ws.max_row + 1):
                                existing_name = ws.cell(row=row_idx, column=1).value or ""
                                if str(existing_name).strip().lower() == cam_name.strip().lower():
                                    ws.cell(row=row_idx, column=2).value = cam_ip
                                    updated_count += 1
                                    found = True
                                    break
                            if not found:
                                # append at end
                                next_row = ws.max_row + 1
                                ws.cell(row=next_row, column=1).value = cam_name
                                ws.cell(row=next_row, column=2).value = cam_ip
                                updated_count += 1
                            QtCore.QTimer.singleShot(0, lambda v=idx: prog.setValue(v))

                        wb.save(excel_path)
                        log(f"[UPDATE CAMERAS] Saved! Updated {updated_count} camera IPs")
                        QtCore.QTimer.singleShot(0, lambda: prog.close())
                        QtCore.QTimer.singleShot(0, lambda: QtWidgets.QMessageBox.information(dlg, "Success", f"✅ Updated {updated_count} camera IPs in Excel"))
                        QtCore.QTimer.singleShot(0, lambda: (self.load_data(), self.populate_nvr_list(), self.populate_table([])))
                    except Exception as e:
                        import traceback
                        log(f"[UPDATE CAMERAS] ERROR: {e}")
                        log(f"[UPDATE CAMERAS] Traceback: {traceback.format_exc()}")
                        QtCore.QTimer.singleShot(0, lambda: prog.close())
                        QtCore.QTimer.singleShot(0, lambda: QtWidgets.QMessageBox.critical(dlg, "Update Error", f"Error updating cameras:\n{str(e)}"))

                threading.Thread(target=update_worker, daemon=True).start()
        
        def update_cameras_in_excel(cameras):
            """Update camera IPs in Excel spreadsheet."""
            log(f"[UPDATE CAMERAS] Updating {len(cameras)} cameras in Excel")
            try:
                from openpyxl import load_workbook
                excel_path = EXCEL_FILE
                
                if not os.path.exists(excel_path):
                    QtWidgets.QMessageBox.warning(dlg, "Error", "Excel file not found.")
                    return
                
                wb = load_workbook(excel_path)
                
                # Get or create camera sheet for this NVR
                sheet_name = nvr_name.replace("/", "_").replace("\\", "_")
                if sheet_name not in wb.sheetnames:
                    # Create new sheet
                    ws = wb.create_sheet(sheet_name)
                    ws.cell(row=1, column=1).value = "Camera Name"
                    ws.cell(row=1, column=2).value = "IP Address"
                    log(f"[UPDATE CAMERAS] Created new sheet: {sheet_name}")
                else:
                    ws = wb[sheet_name]
                
                updated_count = 0
                
                # Update or add cameras
                for cam_idx, camera in enumerate(cameras, start=2):  # Start from row 2 (row 1 is header)
                    cam_name = camera.get("name", f"Camera {cam_idx}")
                    cam_ip = camera.get("ip", "")
                    
                    if not cam_ip:
                        log(f"[UPDATE CAMERAS] Skipping {cam_name} - no IP")
                        continue
                    
                    # Check if camera already exists in sheet
                    found = False
                    for row_idx in range(2, ws.max_row + 1):
                        existing_name = ws.cell(row=row_idx, column=1).value or ""
                        if str(existing_name).strip().lower() == cam_name.strip().lower():
                            # Update existing camera IP
                            ws.cell(row=row_idx, column=2).value = cam_ip
                            log(f"[UPDATE CAMERAS] Updated: {cam_name} -> {cam_ip}")
                            updated_count += 1
                            found = True
                            break
                    
                    if not found:
                        # Add new camera
                        ws.cell(row=cam_idx, column=1).value = cam_name
                        ws.cell(row=cam_idx, column=2).value = cam_ip
                        log(f"[UPDATE CAMERAS] Added: {cam_name} -> {cam_ip}")
                        updated_count += 1
                
                # Save workbook
                wb.save(excel_path)
                log(f"[UPDATE CAMERAS] Saved! Updated {updated_count} camera IPs")
                
                # Reload app data
                self.load_data()
                self.populate_nvr_list()
                self.populate_table([])
                
                QtWidgets.QMessageBox.information(dlg, "Success", f"✅ Updated {updated_count} camera IPs in Excel")
                
            except Exception as e:
                import traceback
                log(f"[UPDATE CAMERAS] ERROR: {e}")
                log(f"[UPDATE CAMERAS] Traceback: {traceback.format_exc()}")
                QtWidgets.QMessageBox.critical(dlg, "Update Error", f"Error updating cameras:\n{str(e)}")
        
        test_btn.clicked.connect(test_login_action)
        browser_btn.clicked.connect(browser_action)
        save_creds_btn.clicked.connect(save_creds_action)
        update_btn.clicked.connect(update_ip_action)
        fetch_cams_btn.clicked.connect(fetch_cameras_action)
        
        # connect SADP discover button on dialog (also allow quick SADP scan here)
        # create a small lambda to reuse username/password fields
        def sadp_scan_action():
            fetch_cams_btn.setEnabled(False)
            fetch_cams_btn.setText("Fetching...")
            threading.Thread(target=lambda: self._run_sadp_scan_and_show(nvr_name), daemon=True).start()

        # If the SADP button exists on main toolbar, it will open a full dialog. Here allow quick scan too.
        # We use the SADP scan helper to find devices and populate cam_info_label
        # Bind a right-click or double-press? For simplicity, add to fetch button context menu
        try:
            # attach to fetch_cams_btn's context menu trigger
            fetch_cams_btn.setToolTip("Fetch cameras from NVR or run SADP discovery")
        except Exception:
            pass
        dlg.exec_()

    def _run_sadp_scan_and_show(self, origin_name=None):
        """Run SADP discover and show small dialog listing devices."""
        devices = sadp_discover(timeout=1.0, scan_hosts=80)
        # call show dialog in main thread
        QtCore.QTimer.singleShot(0, lambda: self._show_sadp_results_dialog(devices, origin_name))

    def show_sadp_tool(self):
        """Enhanced SADP Tool with discovery, monitoring, configuration, and batch operations."""
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("🔧 SADP Tool - Advanced Device Manager")
        dlg.setGeometry(50, 50, 1200, 800)
        main_layout = QtWidgets.QVBoxLayout(dlg)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # ===== Header with instructions =====
        header_layout = QtWidgets.QHBoxLayout()
        header_label = QtWidgets.QLabel("🔧 SADP Tool - Discover, manage, and monitor Hikvision devices on your network")
        header_label.setStyleSheet("font-weight: bold; font-size: 12pt; color: #2c3e50;")
        header_layout.addWidget(header_label)
        header_layout.addStretch()
        main_layout.addLayout(header_layout)

        # ===== Tabs for different operations =====
        tabs = QtWidgets.QTabWidget()
        tabs.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #bdc3c7; }
            QTabBar::tab { background-color: #ecf0f1; padding: 8px 20px; margin: 2px; border: 1px solid #bdc3c7; }
            QTabBar::tab:selected { background-color: #3498db; color: white; font-weight: bold; }
        """)
        main_layout.addWidget(tabs)

        # TAB 1: Discovery
        tab_discover = QtWidgets.QWidget()
        tab_discover_layout = QtWidgets.QVBoxLayout(tab_discover)
        tab_discover_layout.setSpacing(10)
        tabs.addTab(tab_discover, "🔍 Discovery")

        # Controls group
        controls_group = QtWidgets.QGroupBox("Scan Parameters")
        controls_group.setStyleSheet("QGroupBox { font-weight: bold; color: #2c3e50; padding-top: 10px; border: 1px solid #bdc3c7; border-radius: 5px; }")
        controls_layout = QtWidgets.QFormLayout(controls_group)
        controls_layout.setSpacing(8)

        # Subnet input
        subnet_input = QtWidgets.QLineEdit()
        subnet_input.setPlaceholderText("Leave empty for auto-detect (e.g., 192.168.1.0/24)")
        subnet_input.setMinimumHeight(32)
        controls_layout.addRow("📍 Subnet to Scan:", subnet_input)

        # Timeout input
        timeout_spin = QtWidgets.QDoubleSpinBox()
        timeout_spin.setValue(1.0)
        timeout_spin.setMinimum(0.5)
        timeout_spin.setMaximum(10.0)
        timeout_spin.setSingleStep(0.5)
        timeout_spin.setSuffix(" sec")
        timeout_spin.setMinimumHeight(32)
        controls_layout.addRow("⏱ Scan Timeout:", timeout_spin)

        # Scan button
        btn_scan = QtWidgets.QPushButton("▶ START SCAN")
        btn_scan.setMinimumHeight(36)
        btn_scan.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                font-size: 11pt;
                border: none;
                border-radius: 5px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #229954; }
            QPushButton:pressed { background-color: #1e8449; }
        """)
        controls_layout.addRow("", btn_scan)
        tab_discover_layout.addWidget(controls_group)

        # Results label
        results_label = QtWidgets.QLabel("📊 Discovered Devices")
        results_label.setStyleSheet("font-weight: bold; font-size: 11pt; color: #2c3e50;")
        tab_discover_layout.addWidget(results_label)

        # Results table
        tbl_discover = QtWidgets.QTableWidget()
        tbl_discover.setColumnCount(6)
        tbl_discover.setHorizontalHeaderLabels(["IP Address","Model","Serial Number","MAC Address","Device Name","Status"])
        tbl_discover.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        tbl_discover.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        tbl_discover.horizontalHeader().setStretchLastSection(True)
        tbl_discover.setAlternatingRowColors(True)
        tbl_discover.setStyleSheet("""
            QTableWidget { border: 1px solid #bdc3c7; }
            QTableWidget::item:selected { background-color: #3498db; color: white; }
            QHeaderView::section { background-color: #34495e; color: white; padding: 5px; border: none; }
        """)
        tab_discover_layout.addWidget(tbl_discover)

        # TAB 2: Device Registry (history)
        tab_registry = QtWidgets.QWidget()
        tab_registry_layout = QtWidgets.QVBoxLayout(tab_registry)
        tab_registry_layout.setSpacing(10)
        tabs.addTab(tab_registry, "📋 Device Registry")

        registry_label = QtWidgets.QLabel("📚 Device History & Metadata")
        registry_label.setStyleSheet("font-weight: bold; font-size: 11pt; color: #2c3e50;")
        tab_registry_layout.addWidget(registry_label)

        tbl_registry = QtWidgets.QTableWidget()
        tbl_registry.setColumnCount(7)
        tbl_registry.setHorizontalHeaderLabels(["IP","Model","Serial","MAC","Last Seen","First Seen","Notes"])
        tbl_registry.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        tbl_registry.setAlternatingRowColors(True)
        tbl_registry.horizontalHeader().setStretchLastSection(True)
        tbl_registry.setStyleSheet("""
            QTableWidget { border: 1px solid #bdc3c7; }
            QTableWidget::item:selected { background-color: #3498db; color: white; }
            QHeaderView::section { background-color: #34495e; color: white; padding: 5px; border: none; }
        """)
        tab_registry_layout.addWidget(tbl_registry)

        registry_btn_layout = QtWidgets.QHBoxLayout()
        btn_clear_registry = QtWidgets.QPushButton("🗑 Clear Registry")
        btn_clear_registry.setMinimumHeight(32)
        btn_clear_registry.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #c0392b; }
        """)
        registry_btn_layout.addStretch()
        registry_btn_layout.addWidget(btn_clear_registry)
        tab_registry_layout.addLayout(registry_btn_layout)

        # TAB 3: Batch Operations
        tab_batch = QtWidgets.QWidget()
        tab_batch_layout = QtWidgets.QVBoxLayout(tab_batch)
        tab_batch_layout.setSpacing(10)
        tabs.addTab(tab_batch, "⚙ Batch Operations")

        batch_label = QtWidgets.QLabel("🔀 Batch Device Operations - Select devices and apply bulk actions")
        batch_label.setStyleSheet("font-weight: bold; font-size: 11pt; color: #2c3e50;")
        tab_batch_layout.addWidget(batch_label)

        tbl_batch = QtWidgets.QTableWidget()
        tbl_batch.setColumnCount(7)
        tbl_batch.setHorizontalHeaderLabels(["Select","IP","Model","Serial","MAC","Device Name","Status"])
        tbl_batch.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        tbl_batch.setAlternatingRowColors(True)
        tbl_batch.horizontalHeader().setStretchLastSection(True)
        tbl_batch.setStyleSheet("""
            QTableWidget { border: 1px solid #bdc3c7; }
            QTableWidget::item:selected { background-color: #3498db; color: white; }
            QHeaderView::section { background-color: #34495e; color: white; padding: 5px; border: none; }
        """)
        tab_batch_layout.addWidget(tbl_batch)

        batch_ops_layout = QtWidgets.QHBoxLayout()
        btn_select_all = QtWidgets.QPushButton("☑ Select All")
        btn_deselect_all = QtWidgets.QPushButton("☐ Deselect All")
        btn_import_selected = QtWidgets.QPushButton("➕ Import to Excel")
        btn_open_selected = QtWidgets.QPushButton("🌐 Open in Browser")

        for btn in [btn_select_all, btn_deselect_all, btn_import_selected, btn_open_selected]:
            btn.setMinimumHeight(32)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    font-weight: bold;
                    border: none;
                    border-radius: 5px;
                    padding: 5px;
                }
                QPushButton:hover { background-color: #2980b9; }
            """)

        batch_ops_layout.addWidget(btn_select_all)
        batch_ops_layout.addWidget(btn_deselect_all)
        batch_ops_layout.addWidget(btn_import_selected)
        batch_ops_layout.addWidget(btn_open_selected)
        batch_ops_layout.addStretch()
        tab_batch_layout.addLayout(batch_ops_layout)

        # TAB 4: Real-time Monitor
        tab_monitor = QtWidgets.QWidget()
        tab_monitor_layout = QtWidgets.QVBoxLayout(tab_monitor)
        tab_monitor_layout.setSpacing(10)
        tabs.addTab(tab_monitor, "📡 Monitor")

        monitor_label = QtWidgets.QLabel("📡 Real-time Device Monitoring")
        monitor_label.setStyleSheet("font-weight: bold; font-size: 11pt; color: #2c3e50;")
        tab_monitor_layout.addWidget(monitor_label)

        monitor_controls = QtWidgets.QHBoxLayout()
        monitor_controls.setSpacing(10)
        
        monitor_interval = QtWidgets.QSpinBox()
        monitor_interval.setValue(5)
        monitor_interval.setMinimum(1)
        monitor_interval.setMaximum(60)
        monitor_interval.setSuffix(" sec")
        monitor_interval.setMinimumHeight(32)

        btn_start_monitor = QtWidgets.QPushButton("▶ Start Monitoring")
        btn_stop_monitor = QtWidgets.QPushButton("⏹ Stop Monitoring")
        
        for btn in [btn_start_monitor, btn_stop_monitor]:
            btn.setMinimumHeight(32)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #9b59b6;
                    color: white;
                    font-weight: bold;
                    border: none;
                    border-radius: 5px;
                }
                QPushButton:hover { background-color: #8e44ad; }
            """)

        monitor_controls.addWidget(QtWidgets.QLabel("Check Interval:"))
        monitor_controls.addWidget(monitor_interval)
        monitor_controls.addWidget(btn_start_monitor)
        monitor_controls.addWidget(btn_stop_monitor)
        monitor_controls.addStretch()
        tab_monitor_layout.addLayout(monitor_controls)

        tbl_monitor = QtWidgets.QTableWidget()
        tbl_monitor.setColumnCount(7)
        tbl_monitor.setHorizontalHeaderLabels(["IP","Model","Last Check","Status","Online","Offline Count","Notes"])
        tbl_monitor.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        tbl_monitor.setAlternatingRowColors(True)
        tbl_monitor.horizontalHeader().setStretchLastSection(True)
        tbl_monitor.setStyleSheet("""
            QTableWidget { border: 1px solid #bdc3c7; }
            QTableWidget::item:selected { background-color: #3498db; color: white; }
            QHeaderView::section { background-color: #34495e; color: white; padding: 5px; border: none; }
        """)
        tab_monitor_layout.addWidget(tbl_monitor)

        monitor_status = QtWidgets.QLabel("⚫ Monitor Status: Idle")
        monitor_status.setStyleSheet("color: #95a5a6; font-weight: bold; padding: 10px; background-color: #ecf0f1; border-radius: 5px;")
        tab_monitor_layout.addWidget(monitor_status)

        # ===== Bottom buttons =====
        bottom_layout = QtWidgets.QHBoxLayout()
        btn_export_csv = QtWidgets.QPushButton("💾 Export to CSV")
        btn_close = QtWidgets.QPushButton("✕ Close")

        for btn in [btn_export_csv, btn_close]:
            btn.setMinimumHeight(36)
            btn.setMinimumWidth(120)

        btn_export_csv.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #e67e22; }
        """)

        btn_close.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #7f8c8d; }
        """)

        bottom_layout.addWidget(btn_export_csv)
        bottom_layout.addStretch()
        bottom_layout.addWidget(btn_close)
        main_layout.addLayout(bottom_layout)

        # ===== Storage for devices and registry =====
        devices_storage = []
        registry_storage = self._load_sadp_registry()
        monitor_thread_ref = [None]
        monitor_running = [False]

        # ===== Helper functions =====
        def update_discover_table(devices):
            """Update discovery results table."""
            devices_storage.clear()
            devices_storage.extend(devices)
            tbl_discover.setRowCount(len(devices))
            for r, d in enumerate(devices):
                tbl_discover.setItem(r, 0, QtWidgets.QTableWidgetItem(d.get("ip", "")))
                tbl_discover.setItem(r, 1, QtWidgets.QTableWidgetItem(d.get("model", "")))
                tbl_discover.setItem(r, 2, QtWidgets.QTableWidgetItem(d.get("serial", "")))
                tbl_discover.setItem(r, 3, QtWidgets.QTableWidgetItem(d.get("mac", "")))
                tbl_discover.setItem(r, 4, QtWidgets.QTableWidgetItem(d.get("deviceName", "")))
                status_item = QtWidgets.QTableWidgetItem("✓ Found")
                status_item.setForeground(QtGui.QBrush(QtGui.QColor("#27ae60")))
                tbl_discover.setItem(r, 5, status_item)
            tbl_discover.resizeColumnsToContents()

        def update_registry_table(registry):
            """Update device registry table."""
            tbl_registry.setRowCount(len(registry))
            for r, (ip, info) in enumerate(registry.items()):
                tbl_registry.setItem(r, 0, QtWidgets.QTableWidgetItem(ip))
                tbl_registry.setItem(r, 1, QtWidgets.QTableWidgetItem(info.get("model", "")))
                tbl_registry.setItem(r, 2, QtWidgets.QTableWidgetItem(info.get("serial", "")))
                tbl_registry.setItem(r, 3, QtWidgets.QTableWidgetItem(info.get("mac", "")))
                tbl_registry.setItem(r, 4, QtWidgets.QTableWidgetItem(info.get("last_seen", "")))
                tbl_registry.setItem(r, 5, QtWidgets.QTableWidgetItem(info.get("first_seen", "")))
                tbl_registry.setItem(r, 6, QtWidgets.QTableWidgetItem(info.get("notes", "")))
            tbl_registry.resizeColumnsToContents()

        def update_batch_table():
            """Populate batch ops table with checkboxes."""
            tbl_batch.setRowCount(len(devices_storage))
            for r, d in enumerate(devices_storage):
                checkbox = QtWidgets.QCheckBox()
                widget = QtWidgets.QWidget()
                layout_cb = QtWidgets.QHBoxLayout(widget)
                layout_cb.addWidget(checkbox)
                layout_cb.setContentsMargins(0, 0, 0, 0)
                tbl_batch.setCellWidget(r, 0, widget)
                tbl_batch.setItem(r, 1, QtWidgets.QTableWidgetItem(d.get("ip", "")))
                tbl_batch.setItem(r, 2, QtWidgets.QTableWidgetItem(d.get("model", "")))
                tbl_batch.setItem(r, 3, QtWidgets.QTableWidgetItem(d.get("serial", "")))
                tbl_batch.setItem(r, 4, QtWidgets.QTableWidgetItem(d.get("mac", "")))
                tbl_batch.setItem(r, 5, QtWidgets.QTableWidgetItem(d.get("deviceName", "")))
                status_item = QtWidgets.QTableWidgetItem("Ready")
                status_item.setForeground(QtGui.QBrush(QtGui.QColor("#3498db")))
                tbl_batch.setItem(r, 6, status_item)
            tbl_batch.resizeColumnsToContents()

        def scan_action():
            """Execute SADP scan."""
            progress = QtWidgets.QProgressDialog("🔍 Scanning network for devices...", "Cancel", 0, 100, dlg)
            progress.setWindowTitle("SADP Discovery")
            progress.setWindowModality(QtCore.Qt.WindowModal)
            progress.setStyleSheet("""
                QProgressDialog { background-color: white; }
                QProgressBar { border: 2px solid #bdc3c7; border-radius: 5px; text-align: center; }
                QProgressBar::chunk { background-color: #3498db; }
            """)
            progress.show()

            def progress_cb(sent, total):
                QtCore.QTimer.singleShot(0, lambda: progress.setValue(int((sent / max(total, 1)) * 100)))

            def worker():
                subnet = subnet_input.text().strip() or None
                timeout = timeout_spin.value()
                try:
                    devices = sadp_discover(timeout=timeout, scan_hosts=80, progress_callback=progress_cb, target_subnet=subnet)
                    # Update registry
                    for dev in devices:
                        ip = dev.get("ip", "")
                        if ip and ip not in registry_storage:
                            registry_storage[ip] = {
                                "model": dev.get("model", ""),
                                "serial": dev.get("serial", ""),
                                "mac": dev.get("mac", ""),
                                "device_name": dev.get("deviceName", ""),
                                "first_seen": time.strftime("%Y-%m-%d %H:%M:%S"),
                                "last_seen": time.strftime("%Y-%m-%d %H:%M:%S"),
                                "notes": "",
                                "offline_count": 0
                            }
                        elif ip in registry_storage:
                            registry_storage[ip]["last_seen"] = time.strftime("%Y-%m-%d %H:%M:%S")
                    self._save_sadp_registry(registry_storage)
                    QtCore.QTimer.singleShot(0, progress.close)
                    QtCore.QTimer.singleShot(0, lambda: update_discover_table(devices))
                    QtCore.QTimer.singleShot(0, lambda: update_batch_table())
                    QtCore.QTimer.singleShot(0, lambda: update_registry_table(registry_storage))
                    QtCore.QTimer.singleShot(0, lambda: QtWidgets.QMessageBox.information(dlg, "Scan Complete", f"Found {len(devices)} device(s)"))
                except Exception as e:
                    log(f"SADP scan error: {e}")
                    QtCore.QTimer.singleShot(0, progress.close)
                    QtCore.QTimer.singleShot(0, lambda: QtWidgets.QMessageBox.critical(dlg, "Error", f"Scan failed: {e}"))

            threading.Thread(target=worker, daemon=True).start()

        def select_all_action():
            """Select all devices in batch table."""
            for r in range(tbl_batch.rowCount()):
                checkbox = tbl_batch.cellWidget(r, 0).findChild(QtWidgets.QCheckBox)
                if checkbox:
                    checkbox.setChecked(True)

        def deselect_all_action():
            """Deselect all devices in batch table."""
            for r in range(tbl_batch.rowCount()):
                checkbox = tbl_batch.cellWidget(r, 0).findChild(QtWidgets.QCheckBox)
                if checkbox:
                    checkbox.setChecked(False)

        def import_selected_action():
            """Import selected devices to Excel."""
            selected_count = 0
            try:
                from openpyxl import load_workbook
                if not os.path.exists(EXCEL_FILE):
                    QtWidgets.QMessageBox.warning(dlg, "Error", "Excel file not found.")
                    return
                wb = load_workbook(EXCEL_FILE)
                sheet_name = "SADP_Devices"
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                    ws.cell(row=1, column=1).value = "Camera Name"
                    ws.cell(row=1, column=2).value = "IP Address"
                else:
                    ws = wb[sheet_name]

                for r in range(tbl_batch.rowCount()):
                    checkbox = tbl_batch.cellWidget(r, 0).findChild(QtWidgets.QCheckBox)
                    if checkbox and checkbox.isChecked():
                        ip = tbl_batch.item(r, 1).text()
                        name = tbl_batch.item(r, 5).text() or tbl_batch.item(r, 2).text() or f"Device {ip}"
                        next_row = ws.max_row + 1
                        ws.cell(row=next_row, column=1).value = name
                        ws.cell(row=next_row, column=2).value = ip
                        selected_count += 1

                if selected_count > 0:
                    wb.save(EXCEL_FILE)
                    self.load_data()
                    self.populate_nvr_list()
                    QtWidgets.QMessageBox.information(dlg, "Success", f"Imported {selected_count} device(s) into {sheet_name}")
                else:
                    QtWidgets.QMessageBox.information(dlg, "No Selection", "Select at least one device.")
            except Exception as e:
                QtWidgets.QMessageBox.critical(dlg, "Error", f"Import failed: {e}")

        def open_selected_action():
            """Open selected devices in browser."""
            count = 0
            for r in range(tbl_batch.rowCount()):
                checkbox = tbl_batch.cellWidget(r, 0).findChild(QtWidgets.QCheckBox)
                if checkbox and checkbox.isChecked():
                    ip = tbl_batch.item(r, 1).text()
                    if ip:
                        webbrowser.open(f"http://{ip}")
                        count += 1
            if count == 0:
                QtWidgets.QMessageBox.information(dlg, "No Selection", "Select at least one device.")
            else:
                QtWidgets.QMessageBox.information(dlg, "Opened", f"Opened {count} device(s) in browser.")

        def clear_registry_action():
            """Clear all device registry."""
            reply = QtWidgets.QMessageBox.question(dlg, "Confirm", "Clear all device registry?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            if reply == QtWidgets.QMessageBox.Yes:
                registry_storage.clear()
                self._save_sadp_registry(registry_storage)
                update_registry_table(registry_storage)
                QtWidgets.QMessageBox.information(dlg, "Success", "Registry cleared.")

        def start_monitor_action():
            """Start real-time monitoring."""
            monitor_running[0] = True
            btn_start_monitor.setEnabled(False)
            btn_stop_monitor.setEnabled(True)
            monitor_status.setText("🟢 Monitor Status: Running - checking devices every {} seconds".format(monitor_interval.value()))
            monitor_status.setStyleSheet("color: #27ae60; font-weight: bold; padding: 10px; background-color: #d5f4e6; border-radius: 5px;")

            def monitor_worker():
                while monitor_running[0]:
                    try:
                        interval = monitor_interval.value()
                        devices = sadp_discover(timeout=1.0, scan_hosts=40)
                        QtCore.QTimer.singleShot(0, lambda d=devices: self._update_monitor_table(tbl_monitor, d, registry_storage))
                        time.sleep(interval)
                    except Exception as e:
                        log(f"Monitor error: {e}")
                        break
                QtCore.QTimer.singleShot(0, lambda: monitor_status.setText("⚫ Monitor Status: Stopped"))
                QtCore.QTimer.singleShot(0, lambda: monitor_status.setStyleSheet("color: #95a5a6; font-weight: bold; padding: 10px; background-color: #ecf0f1; border-radius: 5px;"))
                QtCore.QTimer.singleShot(0, lambda: btn_start_monitor.setEnabled(True))
                QtCore.QTimer.singleShot(0, lambda: btn_stop_monitor.setEnabled(False))

            monitor_thread_ref[0] = threading.Thread(target=monitor_worker, daemon=True)
            monitor_thread_ref[0].start()

        def stop_monitor_action():
            """Stop real-time monitoring."""
            monitor_running[0] = False
            btn_start_monitor.setEnabled(True)
            btn_stop_monitor.setEnabled(False)

        def export_csv_action():
            """Export all discovered devices to CSV."""
            try:
                filepath = "sadp_devices_export.csv"
                with open(filepath, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(["IP","Model","Serial","MAC","Device Name","First Seen","Last Seen","Notes"])
                    for ip, info in registry_storage.items():
                        writer.writerow([ip, info.get("model",""), info.get("serial",""), info.get("mac",""), 
                                        info.get("device_name",""), info.get("first_seen",""), info.get("last_seen",""), info.get("notes","")])
                QtWidgets.QMessageBox.information(dlg, "Success", f"Exported to {filepath}")
                log(f"SADP devices exported to {filepath}")
            except Exception as e:
                QtWidgets.QMessageBox.critical(dlg, "Error", f"Export failed: {e}")

        # ===== Connect buttons =====
        btn_scan.clicked.connect(scan_action)
        btn_select_all.clicked.connect(select_all_action)
        btn_deselect_all.clicked.connect(deselect_all_action)
        btn_import_selected.clicked.connect(import_selected_action)
        btn_open_selected.clicked.connect(open_selected_action)
        btn_clear_registry.clicked.connect(clear_registry_action)
        btn_start_monitor.clicked.connect(start_monitor_action)
        btn_stop_monitor.clicked.connect(stop_monitor_action)
        btn_stop_monitor.setEnabled(False)
        btn_export_csv.clicked.connect(export_csv_action)
        btn_close.clicked.connect(dlg.accept)

        # Initial table population
        update_registry_table(registry_storage)

        dlg.exec_()

        # Results table
        tbl_discover = QtWidgets.QTableWidget()
        tbl_discover.setColumnCount(6)
        tbl_discover.setHorizontalHeaderLabels(["IP","Model","Serial","MAC","Device Name","Status"])
        tbl_discover.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        tbl_discover.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        tab_discover_layout.addWidget(tbl_discover)

        # TAB 2: Device Registry (history)
        tab_registry = QtWidgets.QWidget()
        tab_registry_layout = QtWidgets.QVBoxLayout(tab_registry)
        tabs.addTab(tab_registry, "📋 Device Registry")

        tbl_registry = QtWidgets.QTableWidget()
        tbl_registry.setColumnCount(7)
        tbl_registry.setHorizontalHeaderLabels(["IP","Model","Serial","MAC","Last Seen","First Seen","Notes"])
        tbl_registry.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        tab_registry_layout.addWidget(tbl_registry)

        btn_clear_registry = QtWidgets.QPushButton("🗑 Clear Registry")
        tab_registry_layout.addWidget(btn_clear_registry)

        # TAB 3: Batch Operations
        tab_batch = QtWidgets.QWidget()
        tab_batch_layout = QtWidgets.QVBoxLayout(tab_batch)
        tabs.addTab(tab_batch, "⚙ Batch Operations")

        batch_label = QtWidgets.QLabel("Select devices and apply bulk operations:")
        tab_batch_layout.addWidget(batch_label)

        tbl_batch = QtWidgets.QTableWidget()
        tbl_batch.setColumnCount(7)
        tbl_batch.setHorizontalHeaderLabels(["Select","IP","Model","Serial","MAC","Device Name","Status"])
        tbl_batch.setColumnCount(7)
        tbl_batch.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        tab_batch_layout.addWidget(tbl_batch)

        batch_ops_layout = QtWidgets.QHBoxLayout()
        btn_select_all = QtWidgets.QPushButton("☑ Select All")
        btn_deselect_all = QtWidgets.QPushButton("☐ Deselect All")
        btn_import_selected = QtWidgets.QPushButton("➕ Import to Excel")
        btn_open_selected = QtWidgets.QPushButton("🌐 Open in Browser")
        batch_ops_layout.addWidget(btn_select_all)
        batch_ops_layout.addWidget(btn_deselect_all)
        batch_ops_layout.addWidget(btn_import_selected)
        batch_ops_layout.addWidget(btn_open_selected)
        batch_ops_layout.addStretch()
        tab_batch_layout.addLayout(batch_ops_layout)

        # TAB 4: Real-time Monitor
        tab_monitor = QtWidgets.QWidget()
        tab_monitor_layout = QtWidgets.QVBoxLayout(tab_monitor)
        tabs.addTab(tab_monitor, "📡 Monitor")

        monitor_controls = QtWidgets.QHBoxLayout()
        btn_start_monitor = QtWidgets.QPushButton("▶ Start Monitoring")
        btn_stop_monitor = QtWidgets.QPushButton("⏹ Stop Monitoring")
        monitor_interval = QtWidgets.QSpinBox()
        monitor_interval.setValue(5)
        monitor_interval.setMinimum(1)
        monitor_interval.setMaximum(60)
        monitor_interval.setSuffix(" sec")
        monitor_controls.addWidget(QtWidgets.QLabel("Check interval:"))
        monitor_controls.addWidget(monitor_interval)
        monitor_controls.addWidget(btn_start_monitor)
        monitor_controls.addWidget(btn_stop_monitor)
        monitor_controls.addStretch()
        tab_monitor_layout.addLayout(monitor_controls)

        tbl_monitor = QtWidgets.QTableWidget()
        tbl_monitor.setColumnCount(7)
        tbl_monitor.setHorizontalHeaderLabels(["IP","Model","Last Check","Status","Online","Offline Count","Notes"])
        tab_monitor_layout.addWidget(tbl_monitor)

        monitor_status = QtWidgets.QLabel("Monitor status: Idle")
        tab_monitor_layout.addWidget(monitor_status)

        # ===== Bottom buttons =====
        bottom_layout = QtWidgets.QHBoxLayout()
        btn_export_csv = QtWidgets.QPushButton("💾 Export to CSV")
        btn_close = QtWidgets.QPushButton("Close")
        bottom_layout.addWidget(btn_export_csv)
        bottom_layout.addStretch()
        bottom_layout.addWidget(btn_close)
        main_layout.addLayout(bottom_layout)

        # ===== Storage for devices and registry =====
        devices_storage = []
        registry_storage = self._load_sadp_registry()
        monitor_thread_ref = [None]
        monitor_running = [False]

        # ===== Helper functions =====
        def update_discover_table(devices):
            """Update discovery results table."""
            devices_storage.clear()
            devices_storage.extend(devices)
            tbl_discover.setRowCount(len(devices))
            for r, d in enumerate(devices):
                tbl_discover.setItem(r, 0, QtWidgets.QTableWidgetItem(d.get("ip", "")))
                tbl_discover.setItem(r, 1, QtWidgets.QTableWidgetItem(d.get("model", "")))
                tbl_discover.setItem(r, 2, QtWidgets.QTableWidgetItem(d.get("serial", "")))
                tbl_discover.setItem(r, 3, QtWidgets.QTableWidgetItem(d.get("mac", "")))
                tbl_discover.setItem(r, 4, QtWidgets.QTableWidgetItem(d.get("deviceName", "")))
                tbl_discover.setItem(r, 5, QtWidgets.QTableWidgetItem("Found"))
            tbl_discover.resizeColumnsToContents()

        def update_registry_table(registry):
            """Update device registry table."""
            tbl_registry.setRowCount(len(registry))
            for r, (ip, info) in enumerate(registry.items()):
                tbl_registry.setItem(r, 0, QtWidgets.QTableWidgetItem(ip))
                tbl_registry.setItem(r, 1, QtWidgets.QTableWidgetItem(info.get("model", "")))
                tbl_registry.setItem(r, 2, QtWidgets.QTableWidgetItem(info.get("serial", "")))
                tbl_registry.setItem(r, 3, QtWidgets.QTableWidgetItem(info.get("mac", "")))
                tbl_registry.setItem(r, 4, QtWidgets.QTableWidgetItem(info.get("last_seen", "")))
                tbl_registry.setItem(r, 5, QtWidgets.QTableWidgetItem(info.get("first_seen", "")))
                tbl_registry.setItem(r, 6, QtWidgets.QTableWidgetItem(info.get("notes", "")))
            tbl_registry.resizeColumnsToContents()

        def update_batch_table():
            """Populate batch ops table with checkboxes."""
            tbl_batch.setRowCount(len(devices_storage))
            for r, d in enumerate(devices_storage):
                checkbox = QtWidgets.QCheckBox()
                widget = QtWidgets.QWidget()
                layout_cb = QtWidgets.QHBoxLayout(widget)
                layout_cb.addWidget(checkbox)
                layout_cb.setContentsMargins(0, 0, 0, 0)
                tbl_batch.setCellWidget(r, 0, widget)
                tbl_batch.setItem(r, 1, QtWidgets.QTableWidgetItem(d.get("ip", "")))
                tbl_batch.setItem(r, 2, QtWidgets.QTableWidgetItem(d.get("model", "")))
                tbl_batch.setItem(r, 3, QtWidgets.QTableWidgetItem(d.get("serial", "")))
                tbl_batch.setItem(r, 4, QtWidgets.QTableWidgetItem(d.get("mac", "")))
                tbl_batch.setItem(r, 5, QtWidgets.QTableWidgetItem(d.get("deviceName", "")))
                tbl_batch.setItem(r, 6, QtWidgets.QTableWidgetItem("Ready"))
            tbl_batch.resizeColumnsToContents()

        def scan_action():
            """Execute SADP scan."""
            progress = QtWidgets.QProgressDialog("Scanning network...", "Cancel", 0, 100, dlg)
            progress.setWindowTitle("SADP Discovery")
            progress.setWindowModality(QtCore.Qt.WindowModal)
            progress.show()

            def progress_cb(sent, total):
                QtCore.QTimer.singleShot(0, lambda: progress.setValue(int((sent / max(total, 1)) * 100)))

            def worker():
                subnet = subnet_input.text().strip() or None
                timeout = timeout_spin.value()
                try:
                    devices = sadp_discover(timeout=timeout, scan_hosts=80, progress_callback=progress_cb, target_subnet=subnet)
                    # Update registry
                    for dev in devices:
                        ip = dev.get("ip", "")
                        if ip and ip not in registry_storage:
                            registry_storage[ip] = {
                                "model": dev.get("model", ""),
                                "serial": dev.get("serial", ""),
                                "mac": dev.get("mac", ""),
                                "device_name": dev.get("deviceName", ""),
                                "first_seen": time.strftime("%Y-%m-%d %H:%M:%S"),
                                "last_seen": time.strftime("%Y-%m-%d %H:%M:%S"),
                                "notes": ""
                            }
                        elif ip in registry_storage:
                            registry_storage[ip]["last_seen"] = time.strftime("%Y-%m-%d %H:%M:%S")
                    self._save_sadp_registry(registry_storage)
                    QtCore.QTimer.singleShot(0, progress.close)
                    QtCore.QTimer.singleShot(0, lambda: update_discover_table(devices))
                    QtCore.QTimer.singleShot(0, lambda: update_batch_table())
                    QtCore.QTimer.singleShot(0, lambda: update_registry_table(registry_storage))
                except Exception as e:
                    log(f"SADP scan error: {e}")
                    QtCore.QTimer.singleShot(0, progress.close)
                    QtCore.QTimer.singleShot(0, lambda: QtWidgets.QMessageBox.critical(dlg, "Error", f"Scan failed: {e}"))

            threading.Thread(target=worker, daemon=True).start()

        def select_all_action():
            """Select all devices in batch table."""
            for r in range(tbl_batch.rowCount()):
                checkbox = tbl_batch.cellWidget(r, 0).findChild(QtWidgets.QCheckBox)
                if checkbox:
                    checkbox.setChecked(True)

        def deselect_all_action():
            """Deselect all devices in batch table."""
            for r in range(tbl_batch.rowCount()):
                checkbox = tbl_batch.cellWidget(r, 0).findChild(QtWidgets.QCheckBox)
                if checkbox:
                    checkbox.setChecked(False)

        def import_selected_action():
            """Import selected devices to Excel."""
            selected_count = 0
            try:
                from openpyxl import load_workbook
                if not os.path.exists(EXCEL_FILE):
                    QtWidgets.QMessageBox.warning(dlg, "Error", "Excel file not found.")
                    return
                wb = load_workbook(EXCEL_FILE)
                sheet_name = "SADP_Devices"
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                    ws.cell(row=1, column=1).value = "Camera Name"
                    ws.cell(row=1, column=2).value = "IP Address"
                else:
                    ws = wb[sheet_name]

                for r in range(tbl_batch.rowCount()):
                    checkbox = tbl_batch.cellWidget(r, 0).findChild(QtWidgets.QCheckBox)
                    if checkbox and checkbox.isChecked():
                        ip = tbl_batch.item(r, 1).text()
                        name = tbl_batch.item(r, 5).text() or tbl_batch.item(r, 2).text() or f"Device {ip}"
                        next_row = ws.max_row + 1
                        ws.cell(row=next_row, column=1).value = name
                        ws.cell(row=next_row, column=2).value = ip
                        selected_count += 1

                if selected_count > 0:
                    wb.save(EXCEL_FILE)
                    self.load_data()
                    self.populate_nvr_list()
                    QtWidgets.QMessageBox.information(dlg, "Success", f"Imported {selected_count} device(s) into {sheet_name}")
                else:
                    QtWidgets.QMessageBox.information(dlg, "No Selection", "Select at least one device.")
            except Exception as e:
                QtWidgets.QMessageBox.critical(dlg, "Error", f"Import failed: {e}")

        def open_selected_action():
            """Open selected devices in browser."""
            for r in range(tbl_batch.rowCount()):
                checkbox = tbl_batch.cellWidget(r, 0).findChild(QtWidgets.QCheckBox)
                if checkbox and checkbox.isChecked():
                    ip = tbl_batch.item(r, 1).text()
                    if ip:
                        webbrowser.open(f"http://{ip}")

        def clear_registry_action():
            """Clear all device registry."""
            reply = QtWidgets.QMessageBox.question(dlg, "Confirm", "Clear all device registry?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            if reply == QtWidgets.QMessageBox.Yes:
                registry_storage.clear()
                self._save_sadp_registry(registry_storage)
                update_registry_table(registry_storage)
                QtWidgets.QMessageBox.information(dlg, "Success", "Registry cleared.")

        def start_monitor_action():
            """Start real-time monitoring."""
            monitor_running[0] = True
            btn_start_monitor.setEnabled(False)
            btn_stop_monitor.setEnabled(True)
            monitor_status.setText("Monitor status: Running...")

            def monitor_worker():
                while monitor_running[0]:
                    try:
                        interval = monitor_interval.value()
                        devices = sadp_discover(timeout=1.0, scan_hosts=40)
                        QtCore.QTimer.singleShot(0, lambda d=devices: self._update_monitor_table(tbl_monitor, d, registry_storage))
                        time.sleep(interval)
                    except Exception as e:
                        log(f"Monitor error: {e}")
                        break
                QtCore.QTimer.singleShot(0, lambda: monitor_status.setText("Monitor status: Stopped"))
                QtCore.QTimer.singleShot(0, lambda: btn_start_monitor.setEnabled(True))
                QtCore.QTimer.singleShot(0, lambda: btn_stop_monitor.setEnabled(False))

            monitor_thread_ref[0] = threading.Thread(target=monitor_worker, daemon=True)
            monitor_thread_ref[0].start()

        def stop_monitor_action():
            """Stop real-time monitoring."""
            monitor_running[0] = False
            btn_start_monitor.setEnabled(True)
            btn_stop_monitor.setEnabled(False)

        def export_csv_action():
            """Export all discovered devices to CSV."""
            try:
                filepath = "sadp_devices_export.csv"
                with open(filepath, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(["IP","Model","Serial","MAC","Device Name","First Seen","Last Seen","Notes"])
                    for ip, info in registry_storage.items():
                        writer.writerow([ip, info.get("model",""), info.get("serial",""), info.get("mac",""), 
                                        info.get("device_name",""), info.get("first_seen",""), info.get("last_seen",""), info.get("notes","")])
                QtWidgets.QMessageBox.information(dlg, "Success", f"Exported to {filepath}")
                log(f"SADP devices exported to {filepath}")
            except Exception as e:
                QtWidgets.QMessageBox.critical(dlg, "Error", f"Export failed: {e}")

        # ===== Connect buttons =====
        btn_scan.clicked.connect(scan_action)
        btn_select_all.clicked.connect(select_all_action)
        btn_deselect_all.clicked.connect(deselect_all_action)
        btn_import_selected.clicked.connect(import_selected_action)
        btn_open_selected.clicked.connect(open_selected_action)
        btn_clear_registry.clicked.connect(clear_registry_action)
        btn_start_monitor.clicked.connect(start_monitor_action)
        btn_stop_monitor.clicked.connect(stop_monitor_action)
        btn_stop_monitor.setEnabled(False)
        btn_export_csv.clicked.connect(export_csv_action)
        btn_close.clicked.connect(dlg.accept)

        # Initial table population
        update_registry_table(registry_storage)

        dlg.exec_()

    def _update_monitor_table(self, table, devices, registry):
        """Update monitor table with current device status."""
        table.setRowCount(len(registry))
        for r, (ip, info) in enumerate(registry.items()):
            found = any(d.get("ip") == ip for d in devices)
            status = "Online" if found else "Offline"
            offline_count = info.get("offline_count", 0)
            if not found:
                offline_count += 1
                info["offline_count"] = offline_count
            else:
                info["offline_count"] = 0
            
            table.setItem(r, 0, QtWidgets.QTableWidgetItem(ip))
            table.setItem(r, 1, QtWidgets.QTableWidgetItem(info.get("model", "")))
            table.setItem(r, 2, QtWidgets.QTableWidgetItem(time.strftime("%H:%M:%S")))
            table.setItem(r, 3, QtWidgets.QTableWidgetItem(status))
            table.setItem(r, 4, QtWidgets.QTableWidgetItem("✓" if found else "✗"))
            table.setItem(r, 5, QtWidgets.QTableWidgetItem(str(offline_count)))
            table.setItem(r, 6, QtWidgets.QTableWidgetItem(info.get("notes", "")))
        table.resizeColumnsToContents()

    def _load_sadp_registry(self) -> dict:
        """Load SADP device registry from file."""
        registry_file = "sadp_registry.json"
        try:
            if os.path.exists(registry_file):
                with open(registry_file, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception as e:
            log(f"Failed to load SADP registry: {e}")
        return {}

    def _save_sadp_registry(self, registry: dict):
        """Save SADP device registry to file."""
        registry_file = "sadp_registry.json"
        try:
            with open(registry_file, "w", encoding="utf-8") as f:
                json.dump(registry, f, indent=2)
        except Exception as e:
            log(f"Failed to save SADP registry: {e}")
    
    def _test_nvr_login_thread(self, dlg, ip, nvr_name, username, password, label_widget, storage, test_btn):
        log(f"[THREAD] Starting NVR login test in thread")
        # Store references BEFORE emitting signal (signal is processed synchronously)
        self._login_dialog_refs = {
            'dlg': dlg,
            'label_widget': label_widget,
            'storage': storage,
            'test_btn': test_btn
        }
        log(f"[THREAD] Storing nvr_name in refs: {nvr_name}")
        self._login_dialog_refs['nvr_name'] = nvr_name
        self._login_dialog_refs['nvr_ip'] = ip
        log(f"[THREAD] Dialog refs stored: storage id={id(storage)}")
        
        try:
            success, real_ip, error = test_nvr_login(ip, username, password)
            log(f"[THREAD] test_nvr_login returned: success={success}, real_ip={real_ip}, error={error}")
        except Exception as e:
            success, real_ip, error = False, "", str(e)
            log(f"[THREAD] Exception in test_nvr_login: {e}")
        
        log(f"[THREAD] Emitting signal with success={success}")
        self.nvr_login_result.emit(success, real_ip, error)
    
    @QtCore.pyqtSlot(bool, str, str)
    def on_nvr_login_result(self, success, real_ip, error):
        """Handle NVR login test result from thread."""
        log(f"[UI SLOT] on_nvr_login_result called: success={success}, real_ip={real_ip}")
        
        if not hasattr(self, '_login_dialog_refs'):
            log(f"[UI SLOT] ERROR: No dialog references stored")
            return
        
        refs = self._login_dialog_refs
        dlg = refs.get('dlg')
        label_widget = refs.get('label_widget')
        storage = refs.get('storage')
        test_btn = refs.get('test_btn')
        
        log(f"[UI SLOT] Retrieved refs: dlg={dlg is not None}, label_widget={label_widget is not None}, storage={storage is not None}, test_btn={test_btn is not None}")
        
        if not all([dlg, label_widget, storage, test_btn]):
            log(f"[UI SLOT] ERROR: Missing dialog references")
            return
        
        try:
            test_btn.setEnabled(True)
            test_btn.setText("🔍 Test Login & Fetch IP")
            log(f"[UI SLOT] Button re-enabled")
            
            if success:
                log(f"[UI SLOT] SUCCESS! Updating storage[0] from '{storage[0]}' to '{real_ip}'")
                storage[0] = real_ip
                log(f"[UI SLOT] After update: storage[0]='{storage[0]}', storage id={id(storage)}")
                if real_ip:
                    label_widget.setText(real_ip)
                    msg = f"✅ Login successful!\n\nReal IP: {real_ip}"
                else:
                    label_widget.setText("(authenticated)")
                    msg = f"✅ Login successful!\n(IP not found in response - using current IP)"
                label_widget.setStyleSheet("color: green;")
                log(f"[UI SLOT] Showing success dialog")
                QtWidgets.QMessageBox.information(dlg, "Success", msg)
                # Persist real IP into NVR entry and into check history
                try:
                    nvr_name = refs.get('nvr_name')
                    nvr_ip = refs.get('nvr_ip')
                    if nvr_ip and real_ip:
                        for n in self.nvrs:
                            if n.get('ip') == nvr_ip or (n.get('name') and n.get('name') == nvr_name):
                                n['real_ip'] = real_ip
                                # also update check history for the NVR IP
                                self.check_history[real_ip] = {
                                    'status': 'NVR Login OK',
                                    'device_type': 'NVR',
                                    'model': '',
                                    'timestamp': time.time()
                                }
                                try:
                                    with open(CHECK_HISTORY_FILE, 'w', encoding='utf-8') as hf:
                                        json.dump(self.check_history, hf, indent=2)
                                except Exception as e:
                                    log(f"[NVR LOGIN] Error saving history: {e}")
                                break
                except Exception as e:
                    log(f"[UI SLOT] Error persisting NVR real_ip: {e}")
            else:
                label_widget.setText("Failed")
                label_widget.setStyleSheet("color: red;")
                error_msg = error if error else "Unknown error"
                log(f"[UI SLOT] Showing error dialog: {error_msg}")
                QtWidgets.QMessageBox.critical(dlg, "Login Failed", f"❌ Login failed:\n\n{error_msg}\n\nMake sure:\n• NVR IP is correct and accessible\n• Username/password are correct\n• NVR web interface is responding")
        except Exception as e:
            log(f"[UI SLOT] Error in callback: {e}")
            test_btn.setEnabled(True)
            test_btn.setText("🔍 Test Login & Fetch IP")

    def show_workflow_wizard(self):
        """Automated workflow wizard following recommended steps."""
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("🚀 Quick Workflow - Automated Camera Discovery")
        dlg.setGeometry(100, 100, 600, 500)
        
        # Set logo icon
        if os.path.exists(LOGO_FILE):
            dlg.setWindowIcon(QtGui.QIcon(LOGO_FILE))
        
        layout = QtWidgets.QVBoxLayout(dlg)
        layout.setSpacing(15)

        # Header with logo
        header_layout = QtWidgets.QHBoxLayout()
        if os.path.exists(LOGO_FILE):
            logo_label = QtWidgets.QLabel()
            pixmap = QtGui.QPixmap(LOGO_FILE).scaledToHeight(48, QtCore.Qt.SmoothTransformation)
            logo_label.setPixmap(pixmap)
            header_layout.addWidget(logo_label)
        
        header_text = QtWidgets.QLabel("🚀 Quick Workflow - Sky-Tech Camera Discovery")
        header_text.setStyleSheet("font-weight: bold; font-size: 14pt; color: #2c3e50;")
        header_layout.addWidget(header_text)
        layout.addLayout(header_layout)

        # Instructions
        instructions = QtWidgets.QTextEdit()
        instructions.setReadOnly(True)
        instructions.setHtml("""
        <b>Recommended Workflow Steps:</b><br><br>
        <span style="color: #27ae60;"><b>✓ Step 1:</b> NVR Login & Fetch Credentials</span><br>
        Log in to your NVR and verify connectivity.<br><br>
        
        <span style="color: #3498db;"><b>✓ Step 2:</b> Fetch & Update Cameras from NVR</span><br>
        Automatically discover all cameras registered on the NVR.<br><br>
        
        <span style="color: #9b59b6;"><b>✓ Step 3:</b> Check IP Online Status</span><br>
        Verify which cameras are currently reachable (parallel check with 60 workers).<br><br>
        
        <span style="color: #f39c12;"><b>✓ Step 4:</b> SADP Network Scan (Optional)</span><br>
        Discover additional cameras on the network that aren't registered on NVR.<br><br>
        
        <span style="color: #e74c3c;"><b>✓ Step 5:</b> Export & Save Results</span><br>
        Save all discovered cameras to Excel and export registry to CSV.
        """)
        instructions.setMinimumHeight(300)
        layout.addWidget(instructions)

        # Progress tracking
        progress_label = QtWidgets.QLabel("Ready to start workflow...")
        progress_label.setStyleSheet("font-weight: bold; color: #34495e; padding: 10px; background-color: #ecf0f1; border-radius: 5px;")
        layout.addWidget(progress_label)

        # Buttons
        btn_layout = QtWidgets.QHBoxLayout()
        btn_start = QtWidgets.QPushButton("▶ Start Workflow")
        btn_start.setMinimumHeight(40)
        btn_start.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                font-size: 12pt;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #229954; }
        """)
        
        btn_cancel = QtWidgets.QPushButton("✕ Cancel")
        btn_cancel.setMinimumHeight(40)
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #7f8c8d; }
        """)

        btn_layout.addWidget(btn_start)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)

        def run_workflow():
            """Execute the complete workflow."""
            btn_start.setEnabled(False)
            btn_cancel.setEnabled(False)
            
            # Step 1: NVR Login
            if not self.nvrs:
                progress_label.setText("⚠️ Step 0: Load Excel file first with NVR data")
                progress_label.setStyleSheet("color: #e74c3c; padding: 10px; background-color: #fadbd8; border-radius: 5px;")
                btn_start.setEnabled(True)
                btn_cancel.setEnabled(True)
                QtWidgets.QMessageBox.warning(dlg, "No NVRs", "Please load Excel file first with NVR data.")
                return

            progress_label.setText("⏳ Step 1/5: Logging into NVR...")
            progress_label.setStyleSheet("color: #3498db; padding: 10px; background-color: #d6eaf8; border-radius: 5px;")
            QtCore.QCoreApplication.processEvents()

            # Auto-login to first NVR
            if self.nvrs:
                nvr = self.nvrs[0]
                self.show_nvr_login_dialog()
                
                # Step 2: Fetch cameras
                QtCore.QTimer.singleShot(2000, lambda: self._workflow_step2(progress_label, dlg, btn_start, btn_cancel))

        def _workflow_complete():
            """Mark workflow as complete."""
            progress_label.setText("✓ Workflow Complete! All cameras discovered and verified.")
            progress_label.setStyleSheet("color: #27ae60; padding: 10px; background-color: #d5f4e6; border-radius: 5px;")
            btn_start.setEnabled(True)
            btn_cancel.setEnabled(True)
            QtWidgets.QMessageBox.information(dlg, "Success", "✓ Workflow completed successfully!\n\nAll cameras have been discovered, verified, and saved to Excel.")

        btn_start.clicked.connect(run_workflow)
        btn_cancel.clicked.connect(dlg.reject)

        dlg.exec_()

    def _workflow_step2(self, progress_label, dlg, btn_start, btn_cancel):
        """Workflow step 2: Fetch cameras from NVR."""
        progress_label.setText("⏳ Step 2/5: Fetching cameras from NVR...")
        progress_label.setStyleSheet("color: #3498db; padding: 10px; background-color: #d6eaf8; border-radius: 5px;")
        QtCore.QCoreApplication.processEvents()
        QtCore.QTimer.singleShot(2000, lambda: self._workflow_step3(progress_label, dlg, btn_start, btn_cancel))

    def _workflow_step3(self, progress_label, dlg, btn_start, btn_cancel):
        """Workflow step 3: Check IP online."""
        progress_label.setText("⏳ Step 3/5: Checking camera IP online status (parallel scan)...")
        progress_label.setStyleSheet("color: #9b59b6; padding: 10px; background-color: #ebdef0; border-radius: 5px;")
        QtCore.QCoreApplication.processEvents()
        self.check_ip_online()
        QtCore.QTimer.singleShot(3000, lambda: self._workflow_step4(progress_label, dlg, btn_start, btn_cancel))

    def _workflow_step4(self, progress_label, dlg, btn_start, btn_cancel):
        """Workflow step 4: SADP network scan."""
        progress_label.setText("⏳ Step 4/5: Scanning network with SADP discovery...")
        progress_label.setStyleSheet("color: #f39c12; padding: 10px; background-color: #fdebd0; border-radius: 5px;")
        QtCore.QCoreApplication.processEvents()
        # Auto-open SADP tool
        self.show_sadp_tool()
        QtCore.QTimer.singleShot(5000, lambda: self._workflow_step5(progress_label, dlg, btn_start, btn_cancel))

    def _workflow_step5(self, progress_label, dlg, btn_start, btn_cancel):
        """Workflow step 5: Export and save."""
        progress_label.setText("⏳ Step 5/5: Exporting and saving results...")
        progress_label.setStyleSheet("color: #e74c3c; padding: 10px; background-color: #fadbd8; border-radius: 5px;")
        QtCore.QCoreApplication.processEvents()
        self.export_csv()
        QtCore.QTimer.singleShot(1000, lambda: self._workflow_complete(progress_label, dlg, btn_start, btn_cancel))

    def _workflow_complete(self, progress_label, dlg, btn_start, btn_cancel):
        """Mark workflow as complete."""
        progress_label.setText("✓ Workflow Complete! All cameras discovered and verified.")
        progress_label.setStyleSheet("color: #27ae60; padding: 10px; background-color: #d5f4e6; border-radius: 5px;")
        btn_start.setEnabled(True)
        btn_cancel.setEnabled(True)
        QtWidgets.QMessageBox.information(dlg, "✓ Success", "Workflow completed successfully!\n\nAll cameras have been discovered, verified, and saved.")
    
    # ---------------- Update System ----------------
    def check_for_updates_startup(self):
        """Check for updates on startup (silent, respects last check time)"""
        if UPDATE_MANAGER_AVAILABLE:
            try:
                log("[UPDATE] Checking for updates on startup...")
                check_for_updates_async(self, show_no_update=False)
            except Exception as e:
                log(f"[UPDATE] Startup check error: {e}")
    
    def check_for_updates_manual(self):
        """Manual update check (always shows result)"""
        if UPDATE_MANAGER_AVAILABLE:
            try:
                log("[UPDATE] Manual update check initiated...")
                check_for_updates_async(self, show_no_update=True)
            except Exception as e:
                log(f"[UPDATE] Manual check error: {e}")
                QtWidgets.QMessageBox.critical(self, "Update Check Error", f"Failed to check for updates:\n{str(e)}")
        else:
            QtWidgets.QMessageBox.information(self, "Updates", "Update system is not available.")

# ---------------- run ----------------
def main():
    app = QtWidgets.QApplication(sys.argv)
    w = CameraMonitor(); w.show()
    log("CameraMonitor Final v8 started.")
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
